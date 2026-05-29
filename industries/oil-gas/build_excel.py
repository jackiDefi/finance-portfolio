#!/usr/bin/env python3
"""
Oil & Gas — simple 4-sheet operating model (Cover / Inputs / Forecast / Valuation).

Fully formula-driven: change any blue input cell on the Inputs sheet and the
Forecast and Valuation sheets recompute, exactly like the interactive page.

Convention:  blue = input,  black = formula,  bold = output / subtotal.

Run:  python3 build_excel.py   ->   writes oil-gas-model.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------------------
# Industry configuration (defaults approximate Shell FY2024)
# ----------------------------------------------------------------------------
CONFIG = {
    "industry":  "Oil & Gas",
    "company":   "Shell plc (LSE: SHEL)",
    "currency":  "USD billion",
    "per_share": "USD",
    "rev_label": "Revenue",
    "da_label":  "DD&A",
    "inputs": [
        # (label, value, kind)  kind: "num" | "pct" | "int"
        ("Revenue (Y0)",                284.0,  "num"),
        ("Revenue growth",              0.01,   "pct"),
        ("EBITDA margin",               0.23,   "pct"),
        ("DD&A / revenue",              0.09,   "pct"),
        ("Capex intensity",             0.075,  "pct"),
        ("Tax rate",                    0.40,   "pct"),
        ("WACC",                        0.085,  "pct"),
        ("Terminal growth",             0.01,   "pct"),
        ("Net debt (Y0)",               39.0,   "num"),
        ("Shares outstanding (m)",      6100,   "int"),
        ("Cost of debt",                0.05,   "pct"),
        ("WC % of incremental revenue", 0.02,   "pct"),
    ],
}

# ----------------------------------------------------------------------------
# Styling
# ----------------------------------------------------------------------------
NAVY   = "1F3A5F"
ACCENT = "C8884A"
BLUE   = "1155CC"   # input font
GREY   = "F2F4F7"

f_title   = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
f_h2      = Font(name="Calibri", size=12, bold=True, color=NAVY)
f_label   = Font(name="Calibri", size=11, color="222222")
f_input   = Font(name="Calibri", size=11, color=BLUE, bold=True)
f_formula = Font(name="Calibri", size=11, color="222222")
f_bold    = Font(name="Calibri", size=11, bold=True, color="111111")
f_unit    = Font(name="Calibri", size=9, italic=True, color="888888")
f_hdr     = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
f_small   = Font(name="Calibri", size=9, color="666666")

fill_title  = PatternFill("solid", fgColor=NAVY)
fill_hdr    = PatternFill("solid", fgColor=ACCENT)
fill_grey   = PatternFill("solid", fgColor=GREY)

thin = Side(style="thin", color="D0D5DD")
border_b = Border(bottom=thin)
center = Alignment(horizontal="center")
right  = Alignment(horizontal="right")
left   = Alignment(horizontal="left")

PCT = "0.0%"
NUM = "#,##0.0"
INT = "#,##0"
MULT = '0.00"x"'

YEARS = 5
YCOLS = [get_column_letter(3 + t) for t in range(YEARS)]  # C..G


def style_pct(c):  c.number_format = PCT
def style_num(c):  c.number_format = NUM


# ----------------------------------------------------------------------------
def build(cfg, out_path):
    wb = Workbook()

    # ===================== COVER =====================
    ws = wb.active
    ws.title = "Cover"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 90

    ws["B2"] = f"{cfg['industry']} — 5-Year Operating Model"
    ws["B2"].font = f_title
    ws["B2"].fill = fill_title
    ws["B2"].alignment = left
    ws.row_dimensions[2].height = 30

    rows = [
        "",
        f"Case study calibration: {cfg['company']}",
        "",
        "MODEL MAP",
        "  Cover       — this sheet: map, conventions, disclaimer",
        "  Inputs      — every driver in one place (blue cells are inputs)",
        "  Forecast    — 5-year income statement, FCFF build, key ratios",
        "  Valuation   — DCF, terminal value, EV → equity → per-share",
        "",
        "CONVENTIONS",
        "  Blue font    = input you can change",
        "  Black font   = formula (do not overwrite)",
        "  Bold         = output / subtotal",
        "",
        "METHOD",
        "  FCFF = NOPAT + DD&A − capex − ΔWorking capital",
        "  NOPAT = EBIT × (1 − tax).  Interest = net debt × cost of debt.",
        "  ΔWC absorbs a % of incremental revenue.",
        "  Terminal value via Gordon growth, discounted at WACC.",
        "",
        "DISCLAIMER",
        "  Default values approximate the case-study company's most recent",
        "  reported year. Balance-sheet starting positions are illustrative.",
        "  Replace blue input cells with audited figures before any real use.",
        "  Analytical template only — not investment advice.",
    ]
    r = 4
    for line in rows:
        ws.cell(row=r, column=2, value=line)
        ws.cell(row=r, column=2).font = f_h2 if line.isupper() and line else f_label
        r += 1

    # ===================== INPUTS =====================
    wi = wb.create_sheet("Inputs")
    wi.sheet_view.showGridLines = False
    wi.column_dimensions["A"].width = 34
    wi.column_dimensions["B"].width = 14
    wi.column_dimensions["C"].width = 16

    wi["A1"] = "Inputs — drivers"
    wi["A1"].font = f_title
    wi["A1"].fill = fill_title
    wi.merge_cells("A1:C1")
    wi.row_dimensions[1].height = 26

    wi["A3"] = "Driver";  wi["B3"] = "Value";  wi["C3"] = "Unit"
    for col in ("A3", "B3", "C3"):
        wi[col].font = f_hdr; wi[col].fill = fill_hdr
    wi["B3"].alignment = center

    units = {"num": cfg["currency"].replace(" billion", " bn"), "pct": "%", "int": "millions"}
    row = 4
    for label, value, kind in cfg["inputs"]:
        wi.cell(row=row, column=1, value=label).font = f_label
        c = wi.cell(row=row, column=2, value=value)
        c.font = f_input
        c.alignment = right
        if kind == "pct":
            c.number_format = PCT
        elif kind == "int":
            c.number_format = INT
        else:
            c.number_format = NUM
        u = "USD bn" if kind == "num" else ("%" if kind == "pct" else "millions")
        wi.cell(row=row, column=3, value=u).font = f_unit
        row += 1

    # input row map (Inputs!B<row>)
    IB = {label: f"Inputs!$B${4 + i}" for i, (label, _, _) in enumerate(cfg["inputs"])}

    rev0  = IB["Revenue (Y0)"]
    grow  = IB["Revenue growth"]
    marg  = IB["EBITDA margin"]
    da    = IB["DD&A / revenue"]
    capx  = IB["Capex intensity"]
    tax   = IB["Tax rate"]
    wacc  = IB["WACC"]
    tg    = IB["Terminal growth"]
    nd    = IB["Net debt (Y0)"]
    sh    = IB["Shares outstanding (m)"]
    cod   = IB["Cost of debt"]
    wcp   = IB["WC % of incremental revenue"]

    # ===================== FORECAST =====================
    wf = wb.create_sheet("Forecast")
    wf.sheet_view.showGridLines = False
    wf.column_dimensions["A"].width = 26
    for col in YCOLS:
        wf.column_dimensions[col].width = 12

    wf["A1"] = f"Forecast — {cfg['currency']}"
    wf["A1"].font = f_title
    wf["A1"].fill = fill_title
    wf.merge_cells(f"A1:{YCOLS[-1]}1")
    wf.row_dimensions[1].height = 26

    wf["A3"] = cfg["currency"]
    wf["A3"].font = f_small
    for t, col in enumerate(YCOLS, start=1):
        c = wf[f"{col}3"]; c.value = f"Y{t}"; c.font = f_hdr; c.fill = fill_hdr; c.alignment = center

    # row indices
    R = {
        "rev": 4, "ebitda": 5, "da": 6, "ebit": 7, "int": 8, "pretax": 9,
        "tax": 10, "ni": 11, "capex": 12, "dwc": 13, "nopat": 14, "fcff": 15,
    }
    labels = {
        "rev": cfg["rev_label"], "ebitda": "EBITDA", "da": cfg["da_label"], "ebit": "EBIT",
        "int": "Interest", "pretax": "Pre-tax profit", "tax": "Tax", "ni": "Net income",
        "capex": "Capex", "dwc": "ΔWorking capital", "nopat": "NOPAT", "fcff": "FCFF",
    }
    bold_rows = {"ebit", "ni", "fcff"}
    for key, ridx in R.items():
        c = wf.cell(row=ridx, column=1, value=labels[key])
        c.font = f_bold if key in bold_rows else f_label

    for t, col in enumerate(YCOLS, start=1):
        prev = YCOLS[t - 2] if t > 1 else None
        rev_prev = f"{prev}{R['rev']}" if prev else rev0
        # Revenue
        wf[f"{col}{R['rev']}"] = (f"={rev0}*(1+{grow})" if t == 1
                                  else f"={prev}{R['rev']}*(1+{grow})")
        wf[f"{col}{R['ebitda']}"] = f"={col}{R['rev']}*{marg}"
        wf[f"{col}{R['da']}"]     = f"={col}{R['rev']}*{da}"
        wf[f"{col}{R['ebit']}"]   = f"={col}{R['ebitda']}-{col}{R['da']}"
        wf[f"{col}{R['int']}"]    = f"={nd}*{cod}"
        wf[f"{col}{R['pretax']}"] = f"={col}{R['ebit']}-{col}{R['int']}"
        wf[f"{col}{R['tax']}"]    = f"=IF({col}{R['pretax']}>0,{col}{R['pretax']}*{tax},0)"
        wf[f"{col}{R['ni']}"]     = f"={col}{R['pretax']}-{col}{R['tax']}"
        wf[f"{col}{R['capex']}"]  = f"={col}{R['rev']}*{capx}"
        wf[f"{col}{R['dwc']}"]    = f"=({col}{R['rev']}-{rev_prev})*{wcp}"
        wf[f"{col}{R['nopat']}"]  = f"={col}{R['ebit']}*(1-{tax})"
        wf[f"{col}{R['fcff']}"]   = f"={col}{R['nopat']}+{col}{R['da']}-{col}{R['capex']}-{col}{R['dwc']}"
        for key, ridx in R.items():
            cell = wf[f"{col}{ridx}"]
            cell.number_format = NUM
            cell.font = f_bold if key in bold_rows else f_formula
            cell.alignment = right

    # ratios
    rr = 17
    wf.cell(row=rr, column=1, value="Key ratios").font = f_h2
    ratio_defs = [
        ("EBITDA margin", lambda col: f"={col}{R['ebitda']}/{col}{R['rev']}", PCT),
        ("FCF / revenue", lambda col: f"={col}{R['fcff']}/{col}{R['rev']}", PCT),
        ("Capex / revenue", lambda col: f"={col}{R['capex']}/{col}{R['rev']}", PCT),
        ("Net debt / EBITDA", lambda col: f"={nd}/{col}{R['ebitda']}", MULT),
    ]
    for i, (lab, fn, fmt) in enumerate(ratio_defs):
        ridx = rr + 1 + i
        wf.cell(row=ridx, column=1, value=lab).font = f_label
        for col in YCOLS:
            c = wf[f"{col}{ridx}"]; c.value = fn(col); c.number_format = fmt
            c.font = f_formula; c.alignment = right

    # ===================== VALUATION =====================
    wv = wb.create_sheet("Valuation")
    wv.sheet_view.showGridLines = False
    wv.column_dimensions["A"].width = 30
    wv.column_dimensions["B"].width = 14
    for col in YCOLS:
        wv.column_dimensions[col].width = 12

    wv["A1"] = "Valuation — DCF"
    wv["A1"].font = f_title
    wv["A1"].fill = fill_title
    wv.merge_cells(f"A1:{YCOLS[-1]}1")
    wv.row_dimensions[1].height = 26

    wv["A3"] = "Discounted cash flow"; wv["A3"].font = f_h2
    for t, col in enumerate(YCOLS, start=1):
        c = wv[f"{col}4"]; c.value = f"Y{t}"; c.font = f_hdr; c.fill = fill_hdr; c.alignment = center

    wv["A5"] = "FCFF"; wv["A5"].font = f_label
    wv["A6"] = "Discount factor"; wv["A6"].font = f_label
    wv["A7"] = "PV of FCFF"; wv["A7"].font = f_bold
    for t, col in enumerate(YCOLS, start=1):
        wv[f"{col}5"] = f"=Forecast!{col}{R['fcff']}"
        wv[f"{col}6"] = f"=1/(1+{wacc})^{t}"
        wv[f"{col}7"] = f"={col}5*{col}6"
        wv[f"{col}5"].number_format = NUM; wv[f"{col}5"].font = f_formula; wv[f"{col}5"].alignment = right
        wv[f"{col}6"].number_format = "0.000"; wv[f"{col}6"].font = f_formula; wv[f"{col}6"].alignment = right
        wv[f"{col}7"].number_format = NUM; wv[f"{col}7"].font = f_bold; wv[f"{col}7"].alignment = right

    last = YCOLS[-1]
    summary = [
        ("Sum of PV of FCFF",        f"=SUM(C7:{last}7)",                                   NUM, True),
        ("Terminal value (Y5)",      f"=Forecast!{last}{R['fcff']}*(1+{tg})/({wacc}-{tg})", NUM, False),
        ("PV of terminal value",     f"=B10/(1+{wacc})^{YEARS}",                            NUM, True),
        ("Enterprise value",         "=B9+B11",                                             NUM, True),
        ("Less: net debt",           f"={nd}",                                              NUM, False),
        ("Equity value",             "=B12-B13",                                            NUM, True),
        ("Shares outstanding (m)",   f"={sh}",                                              INT, False),
        (f"Per-share value ({cfg['per_share']})", "=B14*1000/B15",                          NUM, True),
        ("Implied FY+1 EV/EBITDA",   f"=B12/Forecast!C{R['ebitda']}",                       MULT, True),
    ]
    rstart = 9
    for i, (lab, formula, fmt, is_bold) in enumerate(summary):
        ridx = rstart + i
        a = wv.cell(row=ridx, column=1, value=lab)
        a.font = f_bold if is_bold else f_label
        b = wv.cell(row=ridx, column=2, value=formula)
        b.number_format = fmt
        b.font = f_bold if is_bold else f_formula
        b.alignment = right
        if is_bold:
            a.fill = fill_grey; b.fill = fill_grey

    wb.save(out_path)
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    import os
    here = os.path.dirname(os.path.abspath(__file__))
    build(CONFIG, os.path.join(here, "oil-gas-model.xlsx"))
