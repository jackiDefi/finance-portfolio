"""
Build telecom-model.xlsx — a 16-sheet, formula-driven three-statement
financial model for a stylised integrated telco.

Default values approximate Saudi Telecom Company (Tadawul: 7010);
balance-sheet starting positions are illustrative — they are NOT
actual STC FY2024 figures.

Run from this directory:
    python3 build_excel.py

Convention:
    - Inputs:   blue font   (0563C1)  on the Assumptions sheet
    - Formulas: black font  (000000)
    - Outputs:  bold font   (000000)
    - Headers:  white bold on navy fill
    - Bands:    light-cream fill on totals / key rollup rows
    - Checks:   green if pass, red if fail

Workbook layout (16 sheets):
    1.  Cover          —  title, model map, colour legend, instructions
    2.  Assumptions    —  every driver / starting balance in one place
    3.  Drivers        —  revenue and subscriber build-up by segment
    4.  IS             —  full income statement, Y1-Y5
    5.  BS             —  balance sheet, Y0 opening + Y1-Y5
    6.  CFS            —  cash flow statement, Y1-Y5
    7.  Equity         —  equity rollforward (NI, dividends, OCI)
    8.  Debt           —  debt schedule (gross debt, interest, movements)
    9.  PPE            —  PP&E rollforward (capex, D&A)
    10. WC             —  working capital schedule (AR, inventory, AP)
    11. Tax            —  tax expense schedule
    12. Budget         —  Y1 monthly budget with quarterly subtotals
    13. Scenarios      —  base / upside / downside switcher
    14. Sensitivity    —  two-way data tables on WACC × g and margin × growth
    15. Valuation      —  DCF + multiple cross-check + per-share
    16. Checks         —  BS balance, CFS reconciliation, cash positivity
"""

from datetime import date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName


# =============================================================================
# 1. STYLE PRIMITIVES
# =============================================================================

BLUE     = "0563C1"
NAVY     = "0A1828"
COPPER   = "B87333"
LIGHT_BG = "F1EFE6"
PASS_BG  = "DDF0DC"
FAIL_BG  = "F8D7D5"
PASS_FG  = "2C7A3E"
FAIL_FG  = "B33231"

input_font    = Font(name="Calibri", size=11, color=BLUE,    bold=False)
formula_font  = Font(name="Calibri", size=11, color="000000", bold=False)
output_font   = Font(name="Calibri", size=11, color="000000", bold=True)
header_font   = Font(name="Calibri", size=11, color="FFFFFF", bold=True)
title_font    = Font(name="Calibri", size=18, color=NAVY,    bold=True)
sub_font      = Font(name="Calibri", size=10, color="5A6573", italic=True)
section_font  = Font(name="Calibri", size=10, color=COPPER,  bold=True)
note_font     = Font(name="Calibri", size= 9, color="5A6573")
pass_font     = Font(name="Calibri", size=11, color=PASS_FG, bold=True)
fail_font     = Font(name="Calibri", size=11, color=FAIL_FG, bold=True)

header_fill   = PatternFill("solid", fgColor=NAVY)
band_fill     = PatternFill("solid", fgColor=LIGHT_BG)
pass_fill     = PatternFill("solid", fgColor=PASS_BG)
fail_fill     = PatternFill("solid", fgColor=FAIL_BG)

right  = Alignment(horizontal="right")
left   = Alignment(horizontal="left")
center = Alignment(horizontal="center")

thin   = Side(border_style="thin", color="D4D2C8")
medium = Side(border_style="medium", color="0A1828")
bottom_border    = Border(bottom=thin)
top_border       = Border(top=thin)
top_bottom       = Border(top=thin, bottom=thin)
double_underline = Border(top=thin, bottom=medium)

NUM_BN  = '#,##0.0;[Red]-#,##0.0'
NUM_PCT = '0.0%;[Red]-0.0%'
NUM_X   = '0.00"x"'
NUM_INT = '#,##0'
NUM_DAYS = '#,##0" d"'


# =============================================================================
# 2. DEFAULT VALUES  (mirror the JS DEFAULTS where overlapping)
# =============================================================================

D = {
    # ---- Operating drivers ----
    "rev0":              75.0,     # SAR bn, service revenue Y0
    "growth":            0.06,     # annual %
    "ebitda_target":     0.36,     # display only; actual EBITDA is sum-of-costs
    "da_pct":            0.16,     # D&A / revenue
    "capex_pct":         0.18,     # capex / revenue

    # ---- Cost structure (must sum to 1 - ebitda_target = 0.64) ----
    "cogs_pct":          0.22,     # COGS / interconnect / content
    "network_opex_pct":  0.14,     # network operating costs
    "employee_pct":      0.12,     # employee costs
    "cust_acq_pct":      0.05,     # commercial / customer acquisition
    "other_opex_pct":    0.11,     # G&A and other

    # ---- Other P&L items ----
    "finance_income_pct": 0.005,   # finance income / revenue (treasury yield etc.)

    # ---- Financial assumptions ----
    "tax":               0.15,     # tax / zakat
    "wacc":              0.09,
    "g":                 0.03,     # terminal growth
    "kd":                0.05,     # cost of debt on net debt (interest)
    "div_payout":        0.50,     # payout of net income

    # ---- Working-capital ratios ----
    "dso":               75,       # days sales outstanding
    "dio":               30,       # days inventory outstanding
    "dpo":               90,       # days payables outstanding

    # ---- Balance sheet Y0 (illustrative — NOT actual STC figures) ----
    # Assets
    "cash_y0":           10.0,
    "ar_y0":             15.4,     # = 75 * 75/365
    "inv_y0":             2.5,
    "other_ca_y0":        5.0,
    "ppe_y0":            80.0,
    "goodwill_y0":       25.0,
    "intangibles_y0":    15.0,
    "investments_y0":     8.0,
    "dta_y0":             2.0,
    # Liabilities
    "ap_y0":              7.5,
    "other_cl_y0":       10.0,
    "st_debt_y0":         5.0,
    "lt_debt_y0":        20.0,
    "lease_liab_y0":     18.0,
    "provisions_y0":      5.0,
    "dtl_y0":             3.0,
    "other_ncl_y0":       5.0,
    # Equity (Retained earnings is a plug so total balances)
    "share_capital":      2.0,
    "oci_reserve_y0":     0.0,
    "minority_y0":        7.9,
    # Retained earnings plug: TotalAssets - TotalLiab - SC - OCI - Min
    # = 162.9 - 73.5 - 2 - 0 - 7.9 = 79.5

    # ---- Capital structure ----
    "shares":          5000,       # millions

    # ---- Scenario settings (1 = downside, 2 = base, 3 = upside) ----
    "scen_default":       2,
    "scen_growth":       [0.02, 0.06, 0.09],
    "scen_margin":       [0.33, 0.36, 0.38],
    "scen_capex":        [0.20, 0.18, 0.16],

    # =========================================================================
    # PHASE 1 — Multi-segment revenue & detailed IS
    # =========================================================================

    # ---- Segment 1: Mobile postpaid ----
    "postpaid_subs_y0":    11.0,    # millions
    "postpaid_arpu_y0":   140.0,    # SAR / month
    "postpaid_gross_adds":  0.12,    # gross adds / opening base (annual)
    "postpaid_churn":       0.12,    # annual churn
    "postpaid_arpu_g":      0.020,
    "postpaid_cm":          0.50,    # contribution margin

    # ---- Segment 2: Mobile prepaid ----
    "prepaid_subs_y0":     27.0,
    "prepaid_arpu_y0":      42.0,
    "prepaid_gross_adds":    0.30,
    "prepaid_churn":         0.28,
    "prepaid_arpu_g":        0.010,
    "prepaid_cm":            0.35,

    # ---- Segment 3: Fixed broadband ----
    "fixed_subs_y0":          1.8,
    "fixed_arpu_y0":        420.0,
    "fixed_gross_adds":       0.15,
    "fixed_churn":            0.10,
    "fixed_arpu_g":           0.020,
    "fixed_cm":               0.40,

    # ---- Segments 4-6: modeled directly as revenue (no subs proxy) ----
    "b2b_rev_y0":            18.0,
    "b2b_growth":             0.08,
    "b2b_cm":                 0.45,

    "ict_rev_y0":             8.0,
    "ict_growth":             0.15,
    "ict_cm":                 0.30,

    "wholesale_rev_y0":       8.0,
    "wholesale_growth":       0.02,
    "wholesale_cm":           0.25,

    # ---- Equipment / handset (separate from service revenue) ----
    "equipment_rev_y0":       5.0,
    "equipment_growth":       0.03,
    "equipment_cm":           0.05,

    # ---- Detailed cost ratios (% of total revenue) ----
    # Calibrated so opex sum + equipment COGS ≈ 64% → EBITDA margin ≈ 36%.
    "interconnect_pct":       0.08,
    "content_pct":            0.04,
    "energy_lease_pct":       0.09,
    "other_network_pct":      0.10,
    "commercial_other_pct":   0.04,
    "ga_pct":                 0.07,
    "fx_loss_pct":            0.005,

    # ---- Below-EBITDA IS additions ----
    "intangible_amort_pct":   0.005,  # amortisation of intangibles ex-spectrum (% of revenue)
    "associates_y0":          0.05,   # share of associate profit, SAR bn
    "associates_growth":      0.05,
    "minority_share_ni":      0.05,   # % of group NI attributable to minority

    # ---- Budget seasonality (Q1, Q2, Q3, Q4 monthly weights sum to 1) ----
    # Telecom revenue is fairly flat with slight Q4 uplift from device sales.
    "month_weights": [
        0.0810, 0.0780, 0.0820,    # Q1 = 24.10%
        0.0830, 0.0820, 0.0850,    # Q2 = 25.00%
        0.0820, 0.0830, 0.0840,    # Q3 = 24.90%
        0.0860, 0.0870, 0.0870,    # Q4 = 26.00%
    ],
}

# Derived: starting retained earnings to balance
TOTAL_ASSETS_Y0 = (D["cash_y0"] + D["ar_y0"] + D["inv_y0"] + D["other_ca_y0"]
                   + D["ppe_y0"] + D["goodwill_y0"] + D["intangibles_y0"]
                   + D["investments_y0"] + D["dta_y0"])
TOTAL_LIAB_Y0   = (D["ap_y0"] + D["other_cl_y0"] + D["st_debt_y0"] + D["lt_debt_y0"]
                   + D["lease_liab_y0"] + D["provisions_y0"] + D["dtl_y0"]
                   + D["other_ncl_y0"])
RE_Y0_PLUG = TOTAL_ASSETS_Y0 - TOTAL_LIAB_Y0 - D["share_capital"] - D["oci_reserve_y0"] - D["minority_y0"]
D["re_y0"] = round(RE_Y0_PLUG, 2)


# =============================================================================
# 3. SHEET NAMES
# =============================================================================

SH_COVER  = "Cover"
SH_ASSUM  = "Assumptions"
SH_DRIV   = "Drivers"
SH_SEG    = "Segments"   # Phase 1 — segment P&L
SH_IS     = "IS"
SH_BS     = "BS"
SH_CFS    = "CFS"
SH_EQ     = "Equity"
SH_DEBT   = "Debt"
SH_PPE    = "PPE"
SH_WC     = "WC"
SH_TAX    = "Tax"
SH_BUD    = "Budget"
SH_SCEN   = "Scenarios"
SH_SENS   = "Sensitivity"
SH_VAL    = "Valuation"
SH_CHK    = "Checks"


# =============================================================================
# 4. CELL REGISTRY  (key -> "Sheet!$Col$Row")
# =============================================================================

COORD = {}

def reg(key, sheet, col, row):
    """Register an absolute cross-sheet reference."""
    COORD[key] = f"{sheet}!${col}${row}"
    return COORD[key]

def C(key):
    return COORD[key]

def yc(t):
    """Year column letter. t=0 => B (Y0); t=1..5 => C..G."""
    return get_column_letter(2 + t)


# =============================================================================
# 5. HELPERS
# =============================================================================

def section_header(ws, row, text):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = section_font

def year_header(ws, row, start_t=0, end_t=5, label_y0="Y0 (base)"):
    """Write Y0..Y5 (or subset) header band in row `row`."""
    ws.cell(row=row, column=1).fill = header_fill
    for t in range(start_t, end_t + 1):
        col = 2 + t
        label = label_y0 if t == 0 else f"Y{t}"
        c = ws.cell(row=row, column=col, value=label)
        c.font = header_font
        c.fill = header_fill
        c.alignment = right

def write_label(ws, row, label, *, bold=False, banded=False, indent=0):
    cell = ws.cell(row=row, column=1, value=("  " * indent) + label)
    cell.font = output_font if bold else formula_font
    if banded:
        cell.fill = band_fill
    return cell

def write_input(ws, row, col, value, fmt=NUM_BN, note=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = input_font
    c.number_format = fmt
    c.alignment = right
    if note:
        n = ws.cell(row=row, column=col + 1, value=note)
        n.font = sub_font
    return c

def write_formula(ws, row, col, formula, fmt=NUM_BN, *, bold=False, banded=False):
    c = ws.cell(row=row, column=col, value=formula)
    c.font = output_font if bold else formula_font
    c.number_format = fmt
    c.alignment = right
    if banded:
        c.fill = band_fill
    return c

def fill_row(ws, row, fmt=NUM_BN, start_t=1, end_t=5, formula_for_t=None,
             bold=False, banded=False):
    """Write Y1..Y5 (or subset) formulas using formula_for_t(t) -> str."""
    for t in range(start_t, end_t + 1):
        formula = formula_for_t(t)
        write_formula(ws, row, 2 + t, formula, fmt, bold=bold, banded=banded)


# =============================================================================
# 6. COVER SHEET
# =============================================================================

def build_cover(ws):
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 70

    ws["B2"] = "Telecom & IT — Operating Model"
    ws["B2"].font = title_font
    ws["B3"] = f"Author: Togrul Mirzayev   •   Generated: {date.today().isoformat()}"
    ws["B3"].font = sub_font

    ws["B5"] = "PORTFOLIO"
    ws["B5"].font = section_font
    ws["B6"] = "https://jackidefi.github.io/finance-portfolio/"
    ws["B6"].font = Font(name="Calibri", size=11, color="000000", underline="single")

    ws["B8"] = "MODEL MAP"
    ws["B8"].font = section_font
    sheets = [
        ("Assumptions",  "Every driver and Y0 starting balance — blue cells, change freely."),
        ("Drivers",      "Subscriber and ARPU buildup; service revenue ties back to Assumptions."),
        ("IS",           "Income Statement, fully formula-driven, Y1-Y5."),
        ("BS",           "Balance Sheet, Y0 opening + Y1-Y5 forecast. Must balance."),
        ("CFS",          "Cash Flow Statement, indirect method. Ties to change in cash on BS."),
        ("Equity",       "Equity rollforward — share capital, retained earnings, OCI, minority."),
        ("Debt",         "Debt schedule — gross debt, interest, period movements."),
        ("PPE",          "PP&E rollforward — capex in, D&A out, closing balance to BS."),
        ("WC",           "Working capital schedule — AR, inventory, AP from days ratios."),
        ("Tax",          "Tax expense schedule — current tax, effective rate."),
        ("Budget",       "Y1 broken into 12 months with quarterly subtotals (planning view)."),
        ("Scenarios",    "Downside / Base / Upside switcher; flexes growth, margin, capex."),
        ("Sensitivity",  "Two-way tables: WACC × terminal growth, margin × revenue growth."),
        ("Valuation",    "DCF (Gordon-growth terminal), EV/EBITDA cross-check, per-share."),
        ("Checks",       "BS balance, CFS reconciliation, cash positivity, sources = uses."),
    ]
    for i, (name, desc) in enumerate(sheets):
        ws.cell(row=10 + i, column=2, value=name).font = output_font
        ws.cell(row=10 + i, column=3, value=desc).font = formula_font

    r = 10 + len(sheets) + 2
    ws.cell(row=r, column=2, value="COLOUR LEGEND").font = section_font
    ws.cell(row=r+1, column=2, value="Blue").font = input_font
    ws.cell(row=r+1, column=3, value="Hard-coded input — change freely.").font = formula_font
    ws.cell(row=r+2, column=2, value="Black").font = formula_font
    ws.cell(row=r+2, column=3, value="Formula referencing other cells.").font = formula_font
    ws.cell(row=r+3, column=2, value="Bold").font = output_font
    ws.cell(row=r+3, column=3, value="Subtotal or key model output.").font = formula_font
    ws.cell(row=r+4, column=2, value="Green / red fills").font = formula_font
    ws.cell(row=r+4, column=3, value="Checks sheet — pass / fail status.").font = formula_font

    r += 6
    ws.cell(row=r, column=2, value="DATA DISCLAIMER").font = section_font
    disclaimer = (
        "Default values are calibrated to roughly approximate STC (Tadawul: 7010) at the "
        "consolidated level, but starting balance-sheet line items are illustrative — "
        "they are NOT extracted from actual STC filings. The model demonstrates structure "
        "and methodology. For a publication-grade analysis, replace all blue input cells "
        "with figures verified against STC's audited financial statements."
    )
    cell = ws.cell(row=r+1, column=2, value=disclaimer)
    cell.font = note_font
    cell.alignment = Alignment(horizontal="left", wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r+1, start_column=2, end_row=r+5, end_column=3)
    ws.row_dimensions[r+1].height = 15
    for rr in range(r+1, r+6):
        ws.row_dimensions[rr].height = 15

    r += 7
    ws.cell(row=r, column=2, value="MODEL CONVENTIONS").font = section_font
    conventions = [
        "All figures in SAR billion unless noted (subscribers in millions, ARPU in SAR / month).",
        "Forecast horizon: 5 explicit years (Y1-Y5). Y0 is the base / opening period.",
        "Interest expense uses cost-of-debt × opening total debt to avoid circular references.",
        "Working capital is built from DSO / DIO / DPO ratios applied to revenue and COGS.",
        "Depreciation is driven by D&A as % of revenue; PP&E rolls forward as opening + capex − D&A.",
        "Tax is a single effective rate applied to pre-tax income; deferred tax held flat.",
        "Dividends = payout ratio × net income; equity rolls forward through the Equity sheet.",
        "Terminal value via Gordon growth: FCFF_Y5 × (1+g) ÷ (WACC − g).",
        "BS plugs to cash from the CFS — if every BS movement is correctly mirrored in CFS, the BS balances.",
    ]
    for i, conv in enumerate(conventions):
        c = ws.cell(row=r+1+i, column=2, value=f"•  {conv}")
        c.font = formula_font
        ws.merge_cells(start_row=r+1+i, start_column=2, end_row=r+1+i, end_column=3)


# =============================================================================
# 7. ASSUMPTIONS SHEET
# =============================================================================

def build_assumptions(ws):
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 60

    ws["A1"] = "ASSUMPTIONS"
    ws["A1"].font = section_font
    ws["A2"] = "All inputs to the model live on this sheet. Change blue cells freely."
    ws["A2"].font = sub_font

    rows = [
        # (row, key, label, value, fmt, note)
        # ---- Operating drivers ----
        ( 4, None,              "Operating drivers",                  None,                NUM_BN,  None),
        ( 5, "rev0",            "Service revenue Y0 (SAR bn)",        D["rev0"],           NUM_BN,  "Year 0 base; everything scales from here."),
        ( 6, "growth",          "Revenue growth (annual)",            D["growth"],         NUM_PCT, "Applied to every forecast year (Scenarios sheet can override)."),
        ( 7, "ebitda_target",   "EBITDA margin (target, info only)",  D["ebitda_target"],  NUM_PCT, "Reference value. Actual EBITDA is computed bottom-up from cost lines below."),
        ( 8, "da_pct",          "D&A / revenue",                      D["da_pct"],         NUM_PCT, "Drives PP&E rollforward and depreciation expense."),
        ( 9, "capex_pct",       "Capex intensity",                    D["capex_pct"],      NUM_PCT, "Capex ÷ revenue, fed into PP&E and CFS."),

        # ---- Cost split ----
        (11, None,              "Cost structure (% of revenue)",      None,                NUM_BN,  None),
        (12, "cogs_pct",        "COGS / interconnect / content",      D["cogs_pct"],       NUM_PCT, "Variable; partly indexed to revenue."),
        (13, "network_opex_pct","Network opex",                       D["network_opex_pct"],NUM_PCT,"Site lease, energy, maintenance — mostly fixed."),
        (14, "employee_pct",    "Employee costs",                     D["employee_pct"],   NUM_PCT, "Quasi-fixed."),
        (15, "cust_acq_pct",    "Customer acquisition / commercial",  D["cust_acq_pct"],   NUM_PCT, "Subsidies, dealer commissions."),
        (16, "other_opex_pct",  "Other opex / G&A",                   D["other_opex_pct"], NUM_PCT, "Catch-all; sum of all five lines + EBITDA margin should = 100%."),
        (17, "finance_income_pct","Finance income / revenue",         D["finance_income_pct"],NUM_PCT,"Treasury yield on cash and short-term investments."),

        # ---- Financial ----
        (19, None,              "Financial",                          None,                NUM_BN,  None),
        (20, "tax",             "Tax / zakat rate",                   D["tax"],            NUM_PCT, "Effective rate applied to pre-tax income."),
        (21, "wacc",            "WACC",                               D["wacc"],           NUM_PCT, "Discount rate for DCF."),
        (22, "g",               "Terminal growth (g)",                D["g"],              NUM_PCT, "Long-run nominal growth; must be < WACC."),
        (23, "kd",              "Cost of debt (interest)",            D["kd"],             NUM_PCT, "Applied to opening total debt for period interest."),
        (24, "div_payout",      "Dividend payout (% of NI)",          D["div_payout"],     NUM_PCT, "Used to compute dividends in equity rollforward & CFS."),

        # ---- Working capital ratios ----
        (26, None,              "Working capital (days)",             None,                NUM_BN,  None),
        (27, "dso",             "Days sales outstanding (DSO)",       D["dso"],            NUM_DAYS,"Drives accounts receivable on BS."),
        (28, "dio",             "Days inventory outstanding (DIO)",   D["dio"],            NUM_DAYS,"Drives inventory on BS."),
        (29, "dpo",             "Days payables outstanding (DPO)",    D["dpo"],            NUM_DAYS,"Drives accounts payable on BS."),

        # ---- Capital structure ----
        (31, None,              "Capital structure",                  None,                NUM_BN,  None),
        (32, "shares",          "Shares outstanding (millions)",      D["shares"],         NUM_INT, "Used for per-share value."),

        # ---- Balance sheet Y0 ----
        (34, None,              "Balance sheet Y0 — Assets",          None,                NUM_BN,  None),
        (35, "cash_y0",         "Cash & equivalents",                 D["cash_y0"],        NUM_BN,  "Opening cash; plugs through CFS thereafter."),
        (36, "ar_y0",           "Accounts receivable",                D["ar_y0"],          NUM_BN,  "Calibrated to DSO × revenue."),
        (37, "inv_y0",          "Inventory",                          D["inv_y0"],         NUM_BN,  None),
        (38, "other_ca_y0",     "Other current assets",               D["other_ca_y0"],    NUM_BN,  "Held flat in forecast."),
        (39, "ppe_y0",          "Property, plant & equipment",        D["ppe_y0"],         NUM_BN,  "Net book value; rolls forward via PPE sheet."),
        (40, "goodwill_y0",     "Goodwill",                           D["goodwill_y0"],    NUM_BN,  "Held flat (no impairment in default case)."),
        (41, "intangibles_y0",  "Intangibles (incl. spectrum)",       D["intangibles_y0"], NUM_BN,  "Held flat."),
        (42, "investments_y0",  "Investments / associates",           D["investments_y0"], NUM_BN,  None),
        (43, "dta_y0",          "Deferred tax assets",                D["dta_y0"],         NUM_BN,  "Held flat."),

        (45, None,              "Balance sheet Y0 — Liabilities",     None,                NUM_BN,  None),
        (46, "ap_y0",           "Accounts payable",                   D["ap_y0"],          NUM_BN,  "Calibrated to DPO × COGS."),
        (47, "other_cl_y0",     "Other current liabilities",          D["other_cl_y0"],    NUM_BN,  None),
        (48, "st_debt_y0",      "Short-term debt",                    D["st_debt_y0"],     NUM_BN,  None),
        (49, "lt_debt_y0",      "Long-term debt",                     D["lt_debt_y0"],     NUM_BN,  None),
        (50, "lease_liab_y0",   "Lease liabilities (IFRS 16)",        D["lease_liab_y0"],  NUM_BN,  "Held flat in forecast (simplification)."),
        (51, "provisions_y0",   "Provisions",                         D["provisions_y0"],  NUM_BN,  None),
        (52, "dtl_y0",          "Deferred tax liabilities",           D["dtl_y0"],         NUM_BN,  "Held flat."),
        (53, "other_ncl_y0",    "Other non-current liabilities",      D["other_ncl_y0"],   NUM_BN,  None),

        (55, None,              "Balance sheet Y0 — Equity",          None,                NUM_BN,  None),
        (56, "share_capital",   "Share capital",                      D["share_capital"],  NUM_BN,  "Held flat (no issuance / buyback)."),
        (57, "re_y0",           "Retained earnings",                  D["re_y0"],          NUM_BN,  "Computed plug so total balance sheet balances at Y0."),
        (58, "oci_reserve_y0",  "OCI reserve",                        D["oci_reserve_y0"], NUM_BN,  "Default: nil movement."),
        (59, "minority_y0",     "Minority interest",                  D["minority_y0"],    NUM_BN,  "Held flat."),

        # ---- Scenarios ----
        (61, None,              "Scenario selector",                  None,                NUM_BN,  None),
        (62, "scen_default",    "Active scenario (1=Down, 2=Base, 3=Up)", D["scen_default"], NUM_INT, "Drives the Scenarios sheet."),

        # ---- Phase 1: Segment 1 — Mobile postpaid ----
        (64, None,              "SEGMENT 1 — Mobile postpaid",        None,                NUM_BN,  None),
        (65, "postpaid_subs_y0",   "Subscribers Y0 (millions)",       D["postpaid_subs_y0"],   '#,##0.0', "Opening base."),
        (66, "postpaid_arpu_y0",   "ARPU Y0 (SAR / month)",           D["postpaid_arpu_y0"],   '#,##0',   "Voice + data + content."),
        (67, "postpaid_gross_adds","Gross adds / opening base",       D["postpaid_gross_adds"],NUM_PCT,   "Annual; subscriber acquisition rate."),
        (68, "postpaid_churn",     "Churn (annual)",                  D["postpaid_churn"],     NUM_PCT,   "Annual blended churn."),
        (69, "postpaid_arpu_g",    "ARPU growth (annual)",            D["postpaid_arpu_g"],    NUM_PCT,   None),
        (70, "postpaid_cm",        "Contribution margin",             D["postpaid_cm"],        NUM_PCT,   "Segment EBITDA / segment revenue."),

        # ---- Segment 2 — Mobile prepaid ----
        (72, None,              "SEGMENT 2 — Mobile prepaid",         None,                NUM_BN,  None),
        (73, "prepaid_subs_y0",    "Subscribers Y0 (millions)",       D["prepaid_subs_y0"],   '#,##0.0', None),
        (74, "prepaid_arpu_y0",    "ARPU Y0 (SAR / month)",           D["prepaid_arpu_y0"],   '#,##0',   None),
        (75, "prepaid_gross_adds", "Gross adds / opening base",       D["prepaid_gross_adds"],NUM_PCT,   "Prepaid runs hot — high gross adds, high churn."),
        (76, "prepaid_churn",      "Churn (annual)",                  D["prepaid_churn"],     NUM_PCT,   None),
        (77, "prepaid_arpu_g",     "ARPU growth (annual)",            D["prepaid_arpu_g"],    NUM_PCT,   None),
        (78, "prepaid_cm",         "Contribution margin",             D["prepaid_cm"],        NUM_PCT,   None),

        # ---- Segment 3 — Fixed broadband ----
        (80, None,              "SEGMENT 3 — Fixed broadband",        None,                NUM_BN,  None),
        (81, "fixed_subs_y0",      "Subscribers Y0 (millions)",       D["fixed_subs_y0"],     '#,##0.0', "Households + SMB."),
        (82, "fixed_arpu_y0",      "ARPU Y0 (SAR / month)",           D["fixed_arpu_y0"],     '#,##0',   "Broadband + IPTV + voice bundle."),
        (83, "fixed_gross_adds",   "Gross adds / opening base",       D["fixed_gross_adds"],  NUM_PCT,   None),
        (84, "fixed_churn",        "Churn (annual)",                  D["fixed_churn"],       NUM_PCT,   None),
        (85, "fixed_arpu_g",       "ARPU growth (annual)",            D["fixed_arpu_g"],      NUM_PCT,   None),
        (86, "fixed_cm",           "Contribution margin",             D["fixed_cm"],          NUM_PCT,   None),

        # ---- Segment 4 — B2B connectivity ----
        (88, None,              "SEGMENT 4 — B2B connectivity",       None,                NUM_BN,  None),
        (89, "b2b_rev_y0",         "Revenue Y0 (SAR bn)",             D["b2b_rev_y0"],        NUM_BN,    "Enterprise & government managed connectivity."),
        (90, "b2b_growth",         "Growth (annual)",                 D["b2b_growth"],        NUM_PCT,   None),
        (91, "b2b_cm",             "Contribution margin",             D["b2b_cm"],            NUM_PCT,   None),

        # ---- Segment 5 — ICT / cloud ----
        (93, None,              "SEGMENT 5 — ICT / cloud services",   None,                NUM_BN,  None),
        (94, "ict_rev_y0",         "Revenue Y0 (SAR bn)",             D["ict_rev_y0"],        NUM_BN,    "STC Solutions: cloud, cybersecurity, system integration."),
        (95, "ict_growth",         "Growth (annual)",                 D["ict_growth"],        NUM_PCT,   "Highest segment growth."),
        (96, "ict_cm",             "Contribution margin",             D["ict_cm"],            NUM_PCT,   "Lower margin (heavy COGS in resold hyperscaler capacity)."),

        # ---- Segment 6 — Wholesale ----
        (98, None,              "SEGMENT 6 — Wholesale / carrier",    None,                NUM_BN,  None),
        (99, "wholesale_rev_y0",   "Revenue Y0 (SAR bn)",             D["wholesale_rev_y0"],  NUM_BN,    "Domestic interconnect + international."),
        (100,"wholesale_growth",   "Growth (annual)",                 D["wholesale_growth"],  NUM_PCT,   None),
        (101,"wholesale_cm",       "Contribution margin",             D["wholesale_cm"],      NUM_PCT,   "Margin pressure from declining voice minutes."),

        # ---- Equipment / handset (separate from service revenue) ----
        (103, None,             "EQUIPMENT — handset sales (separate from service)", None,   NUM_BN,  None),
        (104,"equipment_rev_y0",   "Revenue Y0 (SAR bn)",             D["equipment_rev_y0"],  NUM_BN,    None),
        (105,"equipment_growth",   "Growth (annual)",                 D["equipment_growth"],  NUM_PCT,   None),
        (106,"equipment_cm",       "Contribution margin",             D["equipment_cm"],      NUM_PCT,   "Near-breakeven — drives mobile postpaid acquisition."),

        # ---- Detailed cost ratios ----
        (108, None,             "Detailed cost ratios (% of total revenue)", None,           NUM_BN,  None),
        (109,"interconnect_pct",   "Interconnect & roaming",          D["interconnect_pct"],  NUM_PCT,   "Paid to other operators per minute / GB."),
        (110,"content_pct",        "Content / OTT licensing",         D["content_pct"],       NUM_PCT,   "Sports rights, IPTV content."),
        (111,"energy_lease_pct",   "Network energy + site lease",     D["energy_lease_pct"],  NUM_PCT,   "Per-site fixed-ish opex; sensitive to tower lease structure."),
        (112,"other_network_pct",  "Other network opex",              D["other_network_pct"], NUM_PCT,   "Maintenance, transmission, regulatory fees."),
        (113,"commercial_other_pct","Other commercial / marketing",   D["commercial_other_pct"],NUM_PCT, "Brand spend, retention programmes."),
        (114,"ga_pct",             "G&A / overhead",                  D["ga_pct"],            NUM_PCT,   None),
        (115,"fx_loss_pct",        "FX losses (net)",                 D["fx_loss_pct"],       NUM_PCT,   "EGP / KWD / BHD translation drag."),

        # ---- Below-EBITDA items ----
        (117, None,             "Below-EBITDA items",                 None,                   NUM_BN,  None),
        (118,"intangible_amort_pct","Intangibles amortisation",       D["intangible_amort_pct"],NUM_PCT, "Ex-spectrum (spectrum schedule comes in Phase 2)."),
        (119,"associates_y0",      "Share of associate profit Y0",    D["associates_y0"],     NUM_BN,    "Equity-accounted minority stakes."),
        (120,"associates_growth",  "Associates growth",               D["associates_growth"], NUM_PCT,   None),
        (121,"minority_share_ni",  "Minority share of NI",            D["minority_share_ni"], NUM_PCT,   "Vodafone Egypt 49% non-controlling interest."),
    ]

    for row, key, label, value, fmt, note in rows:
        if value is None:
            # Section header row
            section_header(ws, row, label)
        else:
            ws.cell(row=row, column=1, value=label).font = formula_font
            cell = write_input(ws, row, 2, value, fmt, note)
            if key:
                reg(f"a.{key}", SH_ASSUM, "B", row)


# =============================================================================
# 8. DRIVERS SHEET (revenue buildup by segment)
# =============================================================================

def build_drivers(ws):
    """Six-segment revenue buildup with proper subscriber dynamics.

    Subs-based segments (postpaid, prepaid, fixed): opening base + gross adds
    − churn × average base = closing base. ARPU rolls forward at its own rate.
    Direct-revenue segments (B2B, ICT, wholesale): single growth %.
    """
    ws.column_dimensions["A"].width = 38
    for c in range(2, 8):
        ws.column_dimensions[get_column_letter(c)].width = 13

    ws["A1"] = "REVENUE & SUBSCRIBER DRIVERS"
    ws["A1"].font = section_font
    ws["A2"] = "Six service segments + handset revenue. Subs / ARPU / churn dynamics for consumer mobile and fixed; direct-revenue modeling for B2B, ICT and wholesale."
    ws["A2"].font = sub_font

    year_header(ws, 4, 0, 5)
    r = 6

    # ============================================================
    # PART 1 — SUBSCRIBER DYNAMICS (3 segments)
    # ============================================================
    section_header(ws, r, "SUBSCRIBER DYNAMICS  (millions)"); r += 1

    # For each subs segment, build:  opening → gross adds → churn → closing
    subs_segments = [
        ("Mobile postpaid", "postpaid"),
        ("Mobile prepaid",  "prepaid"),
        ("Fixed broadband", "fixed"),
    ]

    closing_rows = {}  # closing subs row by segment key
    arpu_rows = {}

    for label, key in subs_segments:
        # Sub-section header
        write_label(ws, r, label, bold=True); r += 1

        # Opening subs
        write_label(ws, r, "Opening subscribers", indent=1)
        write_formula(ws, r, 2, '""', '@')  # Y0 has no opening
        for t in range(1, 6):
            col = get_column_letter(2 + t)
            if t == 1:
                # Opening Y1 = Y0 closing = a.{key}_subs_y0
                write_formula(ws, r, 2 + t, f"={C(f'a.{key}_subs_y0')}", '#,##0.0')
            else:
                prev_col = get_column_letter(2 + t - 1)
                write_formula(ws, r, 2 + t, f"={prev_col}{r + 3}", '#,##0.0')
        opening_row = r; r += 1

        # Gross adds = % × opening
        write_label(ws, r, "+ Gross adds", indent=1)
        for t in range(1, 6):
            col = get_column_letter(2 + t)
            write_formula(ws, r, 2 + t, f"={C(f'a.{key}_gross_adds')}*{col}{opening_row}", '#,##0.0')
        gross_row = r; r += 1

        # Churn = − churn% × (opening + closing)/2  →  but to avoid circular, use opening
        write_label(ws, r, "− Churn (× opening base)", indent=1)
        for t in range(1, 6):
            col = get_column_letter(2 + t)
            write_formula(ws, r, 2 + t, f"=-{C(f'a.{key}_churn')}*{col}{opening_row}", '#,##0.0')
        churn_row = r; r += 1

        # Closing subs (Y0 = input; Y1+ = opening + gross − churn)
        write_label(ws, r, "Closing subscribers", indent=1, bold=True, banded=True)
        write_formula(ws, r, 2, f"={C(f'a.{key}_subs_y0')}", '#,##0.0', bold=True, banded=True)
        for t in range(1, 6):
            col = get_column_letter(2 + t)
            formula = f"={col}{opening_row}+{col}{gross_row}+{col}{churn_row}"
            write_formula(ws, r, 2 + t, formula, '#,##0.0', bold=True, banded=True)
        closing_rows[key] = r
        for t in range(0, 6):
            col = get_column_letter(2 + t)
            reg(f"dr.{key}_subs_t{t}", SH_DRIV, col, r)
        r += 2

    # ============================================================
    # PART 2 — ARPU ROLLFORWARD
    # ============================================================
    section_header(ws, r, "ARPU  (SAR / month)"); r += 1

    for label, key in subs_segments:
        write_label(ws, r, label, indent=1)
        # Y0 = input ARPU
        write_formula(ws, r, 2, f"={C(f'a.{key}_arpu_y0')}", '#,##0')
        for t in range(1, 6):
            col = get_column_letter(2 + t)
            prev_col = get_column_letter(2 + t - 1)
            write_formula(ws, r, 2 + t, f"={prev_col}{r}*(1+{C(f'a.{key}_arpu_g')})", '#,##0.0')
        arpu_rows[key] = r
        for t in range(0, 6):
            col = get_column_letter(2 + t)
            reg(f"dr.{key}_arpu_t{t}", SH_DRIV, col, r)
        r += 1

    r += 1

    # ============================================================
    # PART 3 — IMPLIED SEGMENT REVENUE
    # ============================================================
    section_header(ws, r, "IMPLIED SEGMENT REVENUE  (SAR bn)"); r += 1

    seg_rev_rows = {}

    for label, key in subs_segments:
        write_label(ws, r, label, indent=1)
        for t in range(0, 6):
            col = get_column_letter(2 + t)
            # subs (mm) × ARPU (SAR/mo) × 12 / 1000 = SAR bn
            # For consumer mobile (postpaid + prepaid) average subs across the year matters,
            # but to avoid circular and keep clean we use closing subs.
            formula = f"={col}{closing_rows[key]}*{col}{arpu_rows[key]}*12/1000"
            write_formula(ws, r, 2 + t, formula, NUM_BN)
            reg(f"dr.{key}_rev_t{t}", SH_DRIV, col, r)
        seg_rev_rows[key] = r
        r += 1

    # Direct-revenue segments
    direct_segments = [
        ("B2B connectivity", "b2b"),
        ("ICT / cloud",      "ict"),
        ("Wholesale",        "wholesale"),
    ]
    for label, key in direct_segments:
        write_label(ws, r, label, indent=1)
        # Y0 from input
        write_formula(ws, r, 2, f"={C(f'a.{key}_rev_y0')}", NUM_BN)
        for t in range(1, 6):
            col = get_column_letter(2 + t)
            prev_col = get_column_letter(2 + t - 1)
            write_formula(ws, r, 2 + t, f"={prev_col}{r}*(1+{C(f'a.{key}_growth')})", NUM_BN)
        seg_rev_rows[key] = r
        for t in range(0, 6):
            col = get_column_letter(2 + t)
            reg(f"dr.{key}_rev_t{t}", SH_DRIV, col, r)
        r += 1

    r += 1

    # ============================================================
    # PART 4 — TOTAL SERVICE REVENUE + EQUIPMENT + TOTAL REVENUE
    # ============================================================
    write_label(ws, r, "TOTAL SERVICE REVENUE", bold=True, banded=True)
    for t in range(0, 6):
        col = get_column_letter(2 + t)
        cells = [f"{col}{seg_rev_rows[k]}" for _, k in subs_segments + direct_segments]
        write_formula(ws, r, 2 + t, "=" + "+".join(cells), NUM_BN, bold=True, banded=True)
        # Maintain backward-compat key: dr.rev_t{t} = total service revenue
        reg(f"dr.rev_t{t}", SH_DRIV, col, r)
        reg(f"dr.service_rev_t{t}", SH_DRIV, col, r)
    r += 1

    # Equipment revenue (separate line)
    write_label(ws, r, "Equipment / handset revenue")
    write_formula(ws, r, 2, f"={C('a.equipment_rev_y0')}", NUM_BN)
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        prev_col = get_column_letter(2 + t - 1)
        write_formula(ws, r, 2 + t, f"={prev_col}{r}*(1+{C('a.equipment_growth')})", NUM_BN)
    for t in range(0, 6):
        col = get_column_letter(2 + t)
        reg(f"dr.equipment_rev_t{t}", SH_DRIV, col, r)
    eq_row = r; r += 1

    # Total revenue
    service_rev_row = r - 2  # the "TOTAL SERVICE REVENUE" row
    write_label(ws, r, "TOTAL REVENUE  (service + equipment)", bold=True, banded=True)
    for t in range(0, 6):
        col = get_column_letter(2 + t)
        formula = f"={col}{service_rev_row}+{col}{eq_row}"
        write_formula(ws, r, 2 + t, formula, NUM_BN, bold=True, banded=True)
        reg(f"dr.total_rev_t{t}", SH_DRIV, col, r)
    r += 2

    # ============================================================
    # PART 5 — OPERATIONAL KPIs
    # ============================================================
    section_header(ws, r, "OPERATIONAL KPIs"); r += 1

    # Total mobile subs (postpaid + prepaid)
    write_label(ws, r, "Total mobile subscribers (mm)", indent=1)
    for t in range(0, 6):
        col = get_column_letter(2 + t)
        formula = f"={col}{closing_rows['postpaid']}+{col}{closing_rows['prepaid']}"
        write_formula(ws, r, 2 + t, formula, '#,##0.0')
    total_mob_row = r; r += 1

    # Postpaid mix
    write_label(ws, r, "Postpaid mix (% of mobile base)", indent=1)
    for t in range(0, 6):
        col = get_column_letter(2 + t)
        formula = f"={col}{closing_rows['postpaid']}/{col}{total_mob_row}"
        write_formula(ws, r, 2 + t, formula, NUM_PCT)
    r += 1

    # Blended mobile ARPU
    write_label(ws, r, "Blended mobile ARPU (SAR / month)", indent=1)
    for t in range(0, 6):
        col = get_column_letter(2 + t)
        pp_rev = f"{col}{seg_rev_rows['postpaid']}"
        pre_rev = f"{col}{seg_rev_rows['prepaid']}"
        total_subs = f"{col}{total_mob_row}"
        formula = f"=({pp_rev}+{pre_rev})*1000/{total_subs}/12"
        write_formula(ws, r, 2 + t, formula, '#,##0')
    r += 1

    # B2B / total mix
    write_label(ws, r, "B2B + ICT mix (% of service rev)", indent=1)
    for t in range(0, 6):
        col = get_column_letter(2 + t)
        b2b = f"{col}{seg_rev_rows['b2b']}"
        ict = f"{col}{seg_rev_rows['ict']}"
        total_serv = f"{col}{service_rev_row}"
        write_formula(ws, r, 2 + t, f"=({b2b}+{ict})/{total_serv}", NUM_PCT)
    r += 1

    # Service revenue growth
    write_label(ws, r, "Service revenue growth (yoy)", indent=1)
    write_formula(ws, r, 2, '""', '@')
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        prev = get_column_letter(2 + t - 1)
        write_formula(ws, r, 2 + t, f"={col}{service_rev_row}/{prev}{service_rev_row}-1", NUM_PCT)


# =============================================================================
# 8B. SEGMENTS P&L  (Phase 1)
# =============================================================================

SEG_ROWS = {}

def build_segments(ws):
    """Per-segment P&L — revenue, direct costs (from contribution margin %),
    and segment contribution. Reconciliation to IS EBITDA appears at the
    bottom (sum of segment contributions vs IS EBITDA = unallocated overhead).
    """
    ws.column_dimensions["A"].width = 34
    for c in range(2, 8):
        ws.column_dimensions[get_column_letter(c)].width = 13

    ws["A1"] = "SEGMENT P&L  (contribution margin view)"
    ws["A1"].font = section_font
    ws["A2"] = ("Each segment carries its own contribution margin (segment revenue × CM%). "
                "Difference vs. IS EBITDA = unallocated group overhead.")
    ws["A2"].font = sub_font

    year_header(ws, 4, 0, 5)
    r = 6

    seg_defs = [
        ("Mobile postpaid",  "postpaid"),
        ("Mobile prepaid",   "prepaid"),
        ("Fixed broadband",  "fixed"),
        ("B2B connectivity", "b2b"),
        ("ICT / cloud",      "ict"),
        ("Wholesale",        "wholesale"),
        ("Equipment",        "equipment"),
    ]

    # ===== Block 1: Revenue =====
    section_header(ws, r, "REVENUE"); r += 1
    rev_rows = {}
    for label, key in seg_defs:
        write_label(ws, r, label, indent=1)
        for t in range(0, 6):
            col = get_column_letter(2 + t)
            # Pull from Drivers
            if key == "equipment":
                src = C(f"dr.equipment_rev_t{t}")
            else:
                src = C(f"dr.{key}_rev_t{t}")
            write_formula(ws, r, 2 + t, f"={src}", NUM_BN)
        rev_rows[key] = r
        r += 1

    # Service revenue subtotal
    write_label(ws, r, "Total service revenue", bold=True)
    for t in range(0, 6):
        col = get_column_letter(2 + t)
        cells = [f"{col}{rev_rows[k]}" for _, k in seg_defs if k != "equipment"]
        write_formula(ws, r, 2 + t, "=" + "+".join(cells), NUM_BN, bold=True)
    SEG_ROWS["service_rev"] = r; r += 1

    write_label(ws, r, "Total revenue (incl. equipment)", bold=True, banded=True)
    for t in range(0, 6):
        col = get_column_letter(2 + t)
        cells = [f"{col}{rev_rows[k]}" for _, k in seg_defs]
        write_formula(ws, r, 2 + t, "=" + "+".join(cells), NUM_BN, bold=True, banded=True)
    SEG_ROWS["total_rev"] = r
    for t in range(0, 6):
        col = get_column_letter(2 + t)
        reg(f"sg.total_rev_t{t}", SH_SEG, col, r)
    r += 2

    # ===== Block 2: Direct costs (segment-level) =====
    section_header(ws, r, "DIRECT COSTS  (= revenue × (1 − CM%))"); r += 1
    cost_rows = {}
    for label, key in seg_defs:
        write_label(ws, r, label, indent=1)
        for t in range(0, 6):
            col = get_column_letter(2 + t)
            rev_cell = f"{col}{rev_rows[key]}"
            formula = f"=-{rev_cell}*(1-{C(f'a.{key}_cm')})"
            write_formula(ws, r, 2 + t, formula, NUM_BN)
        cost_rows[key] = r
        r += 1

    write_label(ws, r, "Total direct costs", bold=True)
    for t in range(0, 6):
        col = get_column_letter(2 + t)
        cells = [f"{col}{cost_rows[k]}" for _, k in seg_defs]
        write_formula(ws, r, 2 + t, "=" + "+".join(cells), NUM_BN, bold=True)
    SEG_ROWS["total_direct"] = r; r += 2

    # ===== Block 3: Segment contribution =====
    section_header(ws, r, "SEGMENT CONTRIBUTION"); r += 1
    contrib_rows = {}
    for label, key in seg_defs:
        write_label(ws, r, label, indent=1)
        for t in range(0, 6):
            col = get_column_letter(2 + t)
            formula = f"={col}{rev_rows[key]}+{col}{cost_rows[key]}"
            write_formula(ws, r, 2 + t, formula, NUM_BN)
        contrib_rows[key] = r
        r += 1

    write_label(ws, r, "Aggregate segment contribution", bold=True, banded=True)
    for t in range(0, 6):
        col = get_column_letter(2 + t)
        cells = [f"{col}{contrib_rows[k]}" for _, k in seg_defs]
        write_formula(ws, r, 2 + t, "=" + "+".join(cells), NUM_BN, bold=True, banded=True)
    SEG_ROWS["agg_contrib"] = r
    for t in range(0, 6):
        col = get_column_letter(2 + t)
        reg(f"sg.agg_contrib_t{t}", SH_SEG, col, r)
    r += 2

    # ===== Block 4: Contribution margin % =====
    section_header(ws, r, "CONTRIBUTION MARGIN %"); r += 1
    for label, key in seg_defs:
        write_label(ws, r, label, indent=1)
        for t in range(0, 6):
            col = get_column_letter(2 + t)
            formula = f"=IF({col}{rev_rows[key]}=0,0,{col}{contrib_rows[key]}/{col}{rev_rows[key]})"
            write_formula(ws, r, 2 + t, formula, NUM_PCT)
        r += 1

    write_label(ws, r, "Blended CM %", bold=True)
    for t in range(0, 6):
        col = get_column_letter(2 + t)
        formula = f"={col}{SEG_ROWS['agg_contrib']}/{col}{SEG_ROWS['total_rev']}"
        write_formula(ws, r, 2 + t, formula, NUM_PCT, bold=True)
    r += 2

    # ===== Block 5: Reconciliation to group EBITDA =====
    section_header(ws, r, "RECONCILIATION TO GROUP EBITDA"); r += 1
    write_label(ws, r, "Aggregate segment contribution")
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        write_formula(ws, r, 2 + t, f"={col}{SEG_ROWS['agg_contrib']}", NUM_BN)
    r += 1

    write_label(ws, r, "− IS EBITDA (group)")
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        write_formula(ws, r, 2 + t, f"=-{C(f'is.ebitda_t{t}')}", NUM_BN)
    r += 1

    write_label(ws, r, "= Unallocated / shared overhead", bold=True, banded=True)
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        prev2 = get_column_letter(2 + t)
        formula = f"={prev2}{r-2}+{prev2}{r-1}"
        write_formula(ws, r, 2 + t, formula, NUM_BN, bold=True, banded=True)
    r += 1


# =============================================================================
# 9. INCOME STATEMENT
# =============================================================================

# We track IS row positions in a dict so other sheets can reference precisely.
IS_ROWS = {}

def build_is(ws):
    ws.column_dimensions["A"].width = 32
    for c in range(2, 8):
        ws.column_dimensions[get_column_letter(c)].width = 13

    ws["A1"] = "INCOME STATEMENT"
    ws["A1"].font = section_font
    ws["A2"] = "All figures in SAR billion. Five-year explicit forecast."
    ws["A2"].font = sub_font

    # Header row: just Y1..Y5
    ws.cell(row=4, column=1).fill = header_fill
    for t in range(1, 6):
        c = ws.cell(row=4, column=2 + t, value=f"Y{t}")
        c.font = header_font; c.fill = header_fill; c.alignment = right
    # Leave column B blank (Y0 not shown on IS; would be historical)

    r = 5

    # Service revenue (from Drivers)
    write_label(ws, r, "Service revenue", bold=True)
    fill_row(ws, r, formula_for_t=lambda t: f"={C(f'dr.rev_t{t}')}", bold=True)
    IS_ROWS["revenue"] = r
    r += 2

    # Costs
    write_label(ws, r, "Cost lines (% of revenue)"); r += 1

    cost_defs = [
        ("COGS / interconnect / content",  "cogs_pct",         "cogs"),
        ("Network operating costs",        "network_opex_pct", "network"),
        ("Employee costs",                 "employee_pct",     "emp"),
        ("Customer acquisition / commercial","cust_acq_pct",   "cust_acq"),
        ("Other opex / G&A",               "other_opex_pct",   "other"),
    ]
    for label, akey, ikey in cost_defs:
        write_label(ws, r, label, indent=1)
        fill_row(ws, r, formula_for_t=lambda t, k=akey: f"=-{C('a.' + k)}*{ws.cell(row=IS_ROWS['revenue'], column=2 + t).coordinate}")
        IS_ROWS[ikey] = r
        r += 1

    # EBITDA
    r += 1
    write_label(ws, r, "EBITDA", bold=True, banded=True)
    cost_rows = [IS_ROWS[k] for _, _, k in cost_defs]
    def ebitda_formula(t):
        col = get_column_letter(2 + t)
        terms = [f"{col}{IS_ROWS['revenue']}"] + [f"{col}{cr}" for cr in cost_rows]
        return "=" + "+".join(terms)
    fill_row(ws, r, formula_for_t=ebitda_formula, bold=True, banded=True)
    IS_ROWS["ebitda"] = r
    r += 2

    # D&A
    write_label(ws, r, "Depreciation & amortisation")
    fill_row(ws, r, formula_for_t=lambda t: f"=-{C('a.da_pct')}*{ws.cell(row=IS_ROWS['revenue'], column=2 + t).coordinate}")
    IS_ROWS["da"] = r
    r += 1

    # EBIT
    r += 1
    write_label(ws, r, "EBIT (operating profit)", bold=True, banded=True)
    fill_row(ws, r, formula_for_t=lambda t: f"={get_column_letter(2+t)}{IS_ROWS['ebitda']}+{get_column_letter(2+t)}{IS_ROWS['da']}", bold=True, banded=True)
    IS_ROWS["ebit"] = r
    r += 2

    # Finance income (% of revenue)
    write_label(ws, r, "Finance income")
    fill_row(ws, r, formula_for_t=lambda t: f"={C('a.finance_income_pct')}*{get_column_letter(2+t)}{IS_ROWS['revenue']}")
    IS_ROWS["fin_inc"] = r
    r += 1

    # Interest expense (negative): from Debt sheet
    write_label(ws, r, "Interest expense")
    fill_row(ws, r, formula_for_t=lambda t: f"=-{SH_DEBT}!{get_column_letter(2+t)}{DEBT_ROWS_PLACEHOLDER}")
    IS_ROWS["interest"] = r
    r += 1

    # Pre-tax income
    r += 1
    write_label(ws, r, "Pre-tax income", bold=True)
    fill_row(ws, r, formula_for_t=lambda t: f"={get_column_letter(2+t)}{IS_ROWS['ebit']}+{get_column_letter(2+t)}{IS_ROWS['fin_inc']}+{get_column_letter(2+t)}{IS_ROWS['interest']}", bold=True)
    IS_ROWS["pretax"] = r
    r += 1

    # Tax (from Tax schedule)
    write_label(ws, r, "Income tax")
    fill_row(ws, r, formula_for_t=lambda t: f"=-{SH_TAX}!{get_column_letter(2+t)}{TAX_ROWS_PLACEHOLDER}")
    IS_ROWS["tax"] = r
    r += 1

    # Net income
    r += 1
    write_label(ws, r, "NET INCOME", bold=True, banded=True)
    fill_row(ws, r, formula_for_t=lambda t: f"={get_column_letter(2+t)}{IS_ROWS['pretax']}+{get_column_letter(2+t)}{IS_ROWS['tax']}", bold=True, banded=True)
    IS_ROWS["ni"] = r
    r += 2

    # Memo: EBITDA margin
    section_header(ws, r, "MEMO"); r += 1
    write_label(ws, r, "EBITDA margin")
    fill_row(ws, r, fmt=NUM_PCT, formula_for_t=lambda t: f"={get_column_letter(2+t)}{IS_ROWS['ebitda']}/{get_column_letter(2+t)}{IS_ROWS['revenue']}")
    IS_ROWS["ebitda_margin"] = r
    r += 1
    write_label(ws, r, "EBIT margin")
    fill_row(ws, r, fmt=NUM_PCT, formula_for_t=lambda t: f"={get_column_letter(2+t)}{IS_ROWS['ebit']}/{get_column_letter(2+t)}{IS_ROWS['revenue']}")
    IS_ROWS["ebit_margin"] = r
    r += 1
    write_label(ws, r, "Net margin")
    fill_row(ws, r, fmt=NUM_PCT, formula_for_t=lambda t: f"={get_column_letter(2+t)}{IS_ROWS['ni']}/{get_column_letter(2+t)}{IS_ROWS['revenue']}")
    IS_ROWS["net_margin"] = r

    # Register cross-sheet references
    for key, row in IS_ROWS.items():
        for t in range(1, 6):
            col = get_column_letter(2 + t)
            reg(f"is.{key}_t{t}", SH_IS, col, row)


# Placeholders — these get patched after Debt/Tax sheets are built
DEBT_ROWS_PLACEHOLDER = 0
TAX_ROWS_PLACEHOLDER = 0
# (We solve this differently — by passing the row number to build_is via globals
# after Debt and Tax are constructed. Simpler approach: build IS LAST among the
# linked sheets. But IS feeds Equity which feeds BS, so order is: Debt -> PPE ->
# WC -> Tax -> IS -> Equity -> BS -> CFS. Let's restructure.)


# =============================================================================
# 10. DEBT SCHEDULE
# =============================================================================

DEBT_ROWS = {}

def build_debt(ws):
    ws.column_dimensions["A"].width = 32
    for c in range(2, 8):
        ws.column_dimensions[get_column_letter(c)].width = 13

    ws["A1"] = "DEBT SCHEDULE"
    ws["A1"].font = section_font
    ws["A2"] = "Total debt = ST + LT + lease liabilities. Held flat unless overridden (simplification)."
    ws["A2"].font = sub_font

    year_header(ws, 4, 0, 5)

    r = 6

    # Opening total debt = Y0 ST + LT + lease
    write_label(ws, r, "Short-term debt")
    write_formula(ws, r, 2, f"={C('a.st_debt_y0')}", NUM_BN)  # Y0
    for t in range(1, 6):
        prev = ws.cell(row=r, column=2 + t - 1).coordinate
        write_formula(ws, r, 2 + t, f"={prev}", NUM_BN)  # held flat
    DEBT_ROWS["st"] = r; r += 1

    write_label(ws, r, "Long-term debt")
    write_formula(ws, r, 2, f"={C('a.lt_debt_y0')}", NUM_BN)
    for t in range(1, 6):
        prev = ws.cell(row=r, column=2 + t - 1).coordinate
        write_formula(ws, r, 2 + t, f"={prev}", NUM_BN)
    DEBT_ROWS["lt"] = r; r += 1

    write_label(ws, r, "Lease liabilities (IFRS 16)")
    write_formula(ws, r, 2, f"={C('a.lease_liab_y0')}", NUM_BN)
    for t in range(1, 6):
        prev = ws.cell(row=r, column=2 + t - 1).coordinate
        write_formula(ws, r, 2 + t, f"={prev}", NUM_BN)
    DEBT_ROWS["lease"] = r; r += 1

    # Total gross debt
    r += 1
    write_label(ws, r, "Gross debt (total)", bold=True, banded=True)
    for t in range(0, 6):
        col = get_column_letter(2 + t)
        formula = f"={col}{DEBT_ROWS['st']}+{col}{DEBT_ROWS['lt']}+{col}{DEBT_ROWS['lease']}"
        write_formula(ws, r, 2 + t, formula, NUM_BN, bold=True, banded=True)
        if t == 0:
            reg(f"d.gross_y0", SH_DEBT, col, r)
        else:
            reg(f"d.gross_t{t}", SH_DEBT, col, r)
    DEBT_ROWS["gross"] = r; r += 2

    # Net debt = gross - cash (cash comes from BS; for now use Y0 directly and let
    # subsequent years pick up cash from BS sheet rows we'll write in build_bs)
    write_label(ws, r, "Less: cash (from BS)")
    # Y0 cash
    write_formula(ws, r, 2, f"=-{C('a.cash_y0')}", NUM_BN)
    # Y1..Y5: pull from BS
    for t in range(1, 6):
        # BS cash row to be defined; we'll reference by name later via key registry
        col = get_column_letter(2 + t)
        write_formula(ws, r, 2 + t, f"=-BS!{col}{BS_PLACEHOLDER_CASH}", NUM_BN)
    DEBT_ROWS["cash_offset"] = r; r += 1

    r += 1
    write_label(ws, r, "Net debt", bold=True, banded=True)
    for t in range(0, 6):
        col = get_column_letter(2 + t)
        formula = f"={col}{DEBT_ROWS['gross']}+{col}{DEBT_ROWS['cash_offset']}"
        write_formula(ws, r, 2 + t, formula, NUM_BN, bold=True, banded=True)
        if t == 0:
            reg("d.net_y0", SH_DEBT, col, r)
        else:
            reg(f"d.net_t{t}", SH_DEBT, col, r)
    DEBT_ROWS["net"] = r; r += 2

    # Interest expense (positive number)
    write_label(ws, r, "Interest expense (= kd × opening gross debt)")
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        prev_col = get_column_letter(2 + t - 1)
        formula = f"={C('a.kd')}*{prev_col}{DEBT_ROWS['gross']}"
        write_formula(ws, r, 2 + t, formula, NUM_BN)
        reg(f"d.interest_t{t}", SH_DEBT, col, r)
    DEBT_ROWS["interest"] = r


BS_PLACEHOLDER_CASH = 0   # patched below by build_bs writing the actual row


# =============================================================================
# 11. PPE SCHEDULE
# =============================================================================

PPE_ROWS = {}

def build_ppe(ws):
    ws.column_dimensions["A"].width = 32
    for c in range(2, 8):
        ws.column_dimensions[get_column_letter(c)].width = 13

    ws["A1"] = "PP&E ROLLFORWARD"
    ws["A1"].font = section_font
    ws["A2"] = "Opening PP&E + capex − depreciation = closing PP&E. Net basis (no gross / accumulated split)."
    ws["A2"].font = sub_font

    year_header(ws, 4, 0, 5)

    r = 6
    # Opening PP&E
    write_label(ws, r, "Opening PP&E")
    write_formula(ws, r, 2, '""', '@')  # Y0 blank
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        prev_col = get_column_letter(2 + t - 1)
        if t == 1:
            write_formula(ws, r, 2 + t, f"={C('a.ppe_y0')}", NUM_BN)
        else:
            write_formula(ws, r, 2 + t, f"={prev_col}{r + 3}", NUM_BN)  # prev closing
    PPE_ROWS["opening"] = r; r += 1

    write_label(ws, r, "+ Capex")
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        formula = f"={C('a.capex_pct')}*{C(f'is.revenue_t{t}')}"
        write_formula(ws, r, 2 + t, formula, NUM_BN)
        reg(f"p.capex_t{t}", SH_PPE, col, r)
    PPE_ROWS["capex"] = r; r += 1

    write_label(ws, r, "− Depreciation & amortisation")
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        formula = f"=-{C('a.da_pct')}*{C(f'is.revenue_t{t}')}"
        write_formula(ws, r, 2 + t, formula, NUM_BN)
        reg(f"p.da_t{t}", SH_PPE, col, r)
    PPE_ROWS["da"] = r; r += 1

    write_label(ws, r, "Closing PP&E", bold=True, banded=True)
    write_formula(ws, r, 2, f"={C('a.ppe_y0')}", NUM_BN, bold=True, banded=True)
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        formula = f"={col}{PPE_ROWS['opening']}+{col}{PPE_ROWS['capex']}+{col}{PPE_ROWS['da']}"
        write_formula(ws, r, 2 + t, formula, NUM_BN, bold=True, banded=True)
        reg(f"p.closing_t{t}", SH_PPE, col, r)
    reg("p.closing_y0", SH_PPE, "B", r)
    PPE_ROWS["closing"] = r


# =============================================================================
# 12. WORKING CAPITAL SCHEDULE
# =============================================================================

WC_ROWS = {}

def build_wc(ws):
    ws.column_dimensions["A"].width = 32
    for c in range(2, 8):
        ws.column_dimensions[get_column_letter(c)].width = 13

    ws["A1"] = "WORKING CAPITAL SCHEDULE"
    ws["A1"].font = section_font
    ws["A2"] = "AR = DSO × revenue / 365. Inventory = DIO × COGS / 365. AP = DPO × COGS / 365."
    ws["A2"].font = sub_font

    year_header(ws, 4, 0, 5)

    r = 6
    # Memo: revenue and COGS used for ratios
    section_header(ws, r, "Reference"); r += 1
    write_label(ws, r, "Revenue")
    write_formula(ws, r, 2, f"={C('a.rev0')}", NUM_BN)
    for t in range(1, 6):
        write_formula(ws, r, 2 + t, f"={C(f'is.revenue_t{t}')}", NUM_BN)
    rev_row = r; r += 1
    write_label(ws, r, "COGS (positive)")
    write_formula(ws, r, 2, f"={C('a.cogs_pct')}*{C('a.rev0')}", NUM_BN)
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        write_formula(ws, r, 2 + t, f"={C('a.cogs_pct')}*{C(f'is.revenue_t{t}')}", NUM_BN)
    cogs_row = r; r += 2

    # Receivables
    section_header(ws, r, "Working capital balances (closing)"); r += 1
    write_label(ws, r, "Accounts receivable")
    # Y0 = input
    write_formula(ws, r, 2, f"={C('a.ar_y0')}", NUM_BN)
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        formula = f"={C('a.dso')}*{col}{rev_row}/365"
        write_formula(ws, r, 2 + t, formula, NUM_BN)
        reg(f"w.ar_t{t}", SH_WC, col, r)
    reg("w.ar_y0", SH_WC, "B", r)
    WC_ROWS["ar"] = r; r += 1

    write_label(ws, r, "Inventory")
    write_formula(ws, r, 2, f"={C('a.inv_y0')}", NUM_BN)
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        formula = f"={C('a.dio')}*{col}{cogs_row}/365"
        write_formula(ws, r, 2 + t, formula, NUM_BN)
        reg(f"w.inv_t{t}", SH_WC, col, r)
    reg("w.inv_y0", SH_WC, "B", r)
    WC_ROWS["inv"] = r; r += 1

    write_label(ws, r, "Accounts payable")
    write_formula(ws, r, 2, f"={C('a.ap_y0')}", NUM_BN)
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        formula = f"={C('a.dpo')}*{col}{cogs_row}/365"
        write_formula(ws, r, 2 + t, formula, NUM_BN)
        reg(f"w.ap_t{t}", SH_WC, col, r)
    reg("w.ap_y0", SH_WC, "B", r)
    WC_ROWS["ap"] = r; r += 2

    # ΔWC for CFS
    section_header(ws, r, "Working capital movements (to CFS)"); r += 1
    write_label(ws, r, "ΔReceivables (incr = use of cash)")
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        prev = get_column_letter(2 + t - 1)
        formula = f"=-({col}{WC_ROWS['ar']}-{prev}{WC_ROWS['ar']})"
        write_formula(ws, r, 2 + t, formula, NUM_BN)
        reg(f"w.dar_t{t}", SH_WC, col, r)
    WC_ROWS["dar"] = r; r += 1

    write_label(ws, r, "ΔInventory (incr = use of cash)")
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        prev = get_column_letter(2 + t - 1)
        formula = f"=-({col}{WC_ROWS['inv']}-{prev}{WC_ROWS['inv']})"
        write_formula(ws, r, 2 + t, formula, NUM_BN)
        reg(f"w.dinv_t{t}", SH_WC, col, r)
    WC_ROWS["dinv"] = r; r += 1

    write_label(ws, r, "ΔPayables (incr = source of cash)")
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        prev = get_column_letter(2 + t - 1)
        formula = f"=({col}{WC_ROWS['ap']}-{prev}{WC_ROWS['ap']})"
        write_formula(ws, r, 2 + t, formula, NUM_BN)
        reg(f"w.dap_t{t}", SH_WC, col, r)
    WC_ROWS["dap"] = r; r += 2

    # Total
    write_label(ws, r, "Total Δ working capital", bold=True, banded=True)
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        formula = f"={col}{WC_ROWS['dar']}+{col}{WC_ROWS['dinv']}+{col}{WC_ROWS['dap']}"
        write_formula(ws, r, 2 + t, formula, NUM_BN, bold=True, banded=True)
        reg(f"w.dwc_t{t}", SH_WC, col, r)
    WC_ROWS["dwc"] = r


# =============================================================================
# 13. TAX SCHEDULE
# =============================================================================

TAX_ROWS = {}

def build_tax(ws):
    ws.column_dimensions["A"].width = 32
    for c in range(2, 8):
        ws.column_dimensions[get_column_letter(c)].width = 13

    ws["A1"] = "TAX SCHEDULE"
    ws["A1"].font = section_font
    ws["A2"] = "Simplified: current tax = max(0, pretax × effective rate). Deferred tax held flat."
    ws["A2"].font = sub_font

    # Header — only Y1-Y5
    ws.cell(row=4, column=1).fill = header_fill
    for t in range(1, 6):
        c = ws.cell(row=4, column=2 + t, value=f"Y{t}")
        c.font = header_font; c.fill = header_fill; c.alignment = right

    r = 6
    write_label(ws, r, "Pre-tax income (from IS)")
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        write_formula(ws, r, 2 + t, f"={C(f'is.pretax_t{t}')}", NUM_BN)
    TAX_ROWS["pretax"] = r; r += 1

    write_label(ws, r, "Effective tax rate")
    for t in range(1, 6):
        write_formula(ws, r, 2 + t, f"={C('a.tax')}", NUM_PCT)
    TAX_ROWS["rate"] = r; r += 1

    write_label(ws, r, "Current tax expense", bold=True, banded=True)
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        formula = f"=MAX(0,{col}{TAX_ROWS['pretax']}*{col}{TAX_ROWS['rate']})"
        write_formula(ws, r, 2 + t, formula, NUM_BN, bold=True, banded=True)
        reg(f"t.expense_t{t}", SH_TAX, col, r)
    TAX_ROWS["expense"] = r


# =============================================================================
# 14. INCOME STATEMENT — rewritten to use real Debt/Tax row references
# =============================================================================

def build_is_v2(ws):
    """Detailed P&L — segment revenue breakdown, 10-line cost split, full
    finance items, attributable NI to equity / minority, basic EPS, margin memos.
    Layout is deterministic so other sheets can reference cells before this is built
    (see PRE_IS_LAYOUT in main)."""
    ws.column_dimensions["A"].width = 42
    for c in range(2, 8):
        ws.column_dimensions[get_column_letter(c)].width = 13

    ws["A1"] = "INCOME STATEMENT"
    ws["A1"].font = section_font
    ws["A2"] = ("Segment-disclosed revenue + 10-line cost split + finance items + "
                "tax + attributable NI + EPS. All figures SAR billion, EPS in SAR.")
    ws["A2"].font = sub_font

    ws.cell(row=4, column=1).fill = header_fill
    for t in range(1, 6):
        c = ws.cell(row=4, column=2 + t, value=f"Y{t}")
        c.font = header_font; c.fill = header_fill; c.alignment = right

    # ============================================================
    # SECTION 1 — REVENUE (rows 5-14)
    # ============================================================
    section_header(ws, 5, "REVENUE")

    seg_defs = [
        ("Mobile postpaid",  "postpaid"),
        ("Mobile prepaid",   "prepaid"),
        ("Fixed broadband",  "fixed"),
        ("B2B connectivity", "b2b"),
        ("ICT / cloud",      "ict"),
        ("Wholesale",        "wholesale"),
    ]
    for i, (label, key) in enumerate(seg_defs):
        r = 6 + i
        write_label(ws, r, label, indent=1)
        for t in range(1, 6):
            col = get_column_letter(2 + t)
            write_formula(ws, r, 2 + t, f"={C(f'dr.{key}_rev_t{t}')}", NUM_BN)
            reg(f"is.seg_{key}_t{t}", SH_IS, col, r)
        IS_ROWS[f"seg_{key}"] = r

    # Total service revenue at row 12
    r = 12
    write_label(ws, r, "Total service revenue", bold=True)
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        cells = [f"{col}{IS_ROWS[f'seg_{k}']}" for _, k in seg_defs]
        write_formula(ws, r, 2 + t, "=" + "+".join(cells), NUM_BN, bold=True)
        reg(f"is.service_revenue_t{t}", SH_IS, col, r)
    IS_ROWS["service_revenue"] = r

    # Equipment revenue at row 13
    r = 13
    write_label(ws, r, "Equipment / handset revenue")
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        write_formula(ws, r, 2 + t, f"={C(f'dr.equipment_rev_t{t}')}", NUM_BN)
        reg(f"is.equipment_revenue_t{t}", SH_IS, col, r)
    IS_ROWS["equipment_revenue"] = r

    # TOTAL REVENUE at row 14
    r = 14
    write_label(ws, r, "TOTAL REVENUE", bold=True, banded=True)
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        formula = f"={col}{IS_ROWS['service_revenue']}+{col}{IS_ROWS['equipment_revenue']}"
        write_formula(ws, r, 2 + t, formula, NUM_BN, bold=True, banded=True)
        reg(f"is.revenue_t{t}", SH_IS, col, r)
        reg(f"is.total_revenue_t{t}", SH_IS, col, r)
    IS_ROWS["revenue"] = r
    IS_ROWS["total_revenue"] = r

    # ============================================================
    # SECTION 2 — OPERATING COSTS (rows 16-27)
    # ============================================================
    section_header(ws, 16, "OPERATING COSTS")

    cost_lines = [
        # (label, assumption_key OR special, internal_key)
        (17, "Interconnect & roaming",        "interconnect_pct",     "interconnect"),
        (18, "Content / OTT licensing",       "content_pct",          "content"),
        (19, "Equipment COGS",                "_EQ_COGS_",            "equipment_cogs"),
        (20, "Network energy + site lease",   "energy_lease_pct",     "energy_lease"),
        (21, "Other network opex",            "other_network_pct",    "other_network"),
        (22, "Employee costs",                "employee_pct",         "employee"),
        (23, "Customer acquisition (commissions, subsidies)", "cust_acq_pct", "cust_acq"),
        (24, "Other commercial / marketing",  "commercial_other_pct", "commercial_other"),
        (25, "G&A / overhead",                "ga_pct",               "ga"),
        (26, "FX losses (net)",               "fx_loss_pct",          "fx"),
    ]
    for row, label, akey, ikey in cost_lines:
        write_label(ws, row, label, indent=1)
        for t in range(1, 6):
            col = get_column_letter(2 + t)
            if akey == "_EQ_COGS_":
                # = -equipment revenue × (1 - equipment_cm)
                eq_rev = f"{col}{IS_ROWS['equipment_revenue']}"
                formula = f"=-{eq_rev}*(1-{C('a.equipment_cm')})"
            else:
                tot_rev = f"{col}{IS_ROWS['revenue']}"
                formula = f"=-{C('a.' + akey)}*{tot_rev}"
            write_formula(ws, row, 2 + t, formula, NUM_BN)
            reg(f"is.{ikey}_t{t}", SH_IS, col, row)
        IS_ROWS[ikey] = row

    # Total operating costs at row 27
    r = 27
    write_label(ws, r, "Total operating costs", bold=True)
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        cells = [f"{col}{IS_ROWS[k]}" for _, _, _, k in cost_lines]
        write_formula(ws, r, 2 + t, "=" + "+".join(cells), NUM_BN, bold=True)
    IS_ROWS["total_opex"] = r

    # EBITDA at row 28
    r = 28
    write_label(ws, r, "EBITDA", bold=True, banded=True)
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        formula = f"={col}{IS_ROWS['revenue']}+{col}{IS_ROWS['total_opex']}"
        write_formula(ws, r, 2 + t, formula, NUM_BN, bold=True, banded=True)
        reg(f"is.ebitda_t{t}", SH_IS, col, r)
    IS_ROWS["ebitda"] = r

    # ============================================================
    # SECTION 3 — DEPRECIATION & AMORTISATION (rows 30-32)
    # ============================================================
    r = 30
    write_label(ws, r, "Depreciation")
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        write_formula(ws, r, 2 + t, f"={C(f'p.da_t{t}')}", NUM_BN)
        reg(f"is.depreciation_t{t}", SH_IS, col, r)
        reg(f"is.da_t{t}", SH_IS, col, r)
    IS_ROWS["depreciation"] = r
    IS_ROWS["da"] = r

    r = 31
    write_label(ws, r, "Amortisation (intangibles ex-spectrum)")
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        formula = f"=-{C('a.intangible_amort_pct')}*{col}{IS_ROWS['revenue']}"
        write_formula(ws, r, 2 + t, formula, NUM_BN)
        reg(f"is.amortisation_t{t}", SH_IS, col, r)
    IS_ROWS["amortisation"] = r

    r = 32
    write_label(ws, r, "EBIT", bold=True, banded=True)
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        cells = [IS_ROWS["ebitda"], IS_ROWS["depreciation"], IS_ROWS["amortisation"]]
        formula = "=" + "+".join(f"{col}{x}" for x in cells)
        write_formula(ws, r, 2 + t, formula, NUM_BN, bold=True, banded=True)
        reg(f"is.ebit_t{t}", SH_IS, col, r)
    IS_ROWS["ebit"] = r

    # ============================================================
    # SECTION 4 — FINANCE ITEMS & PBT (rows 34-37)
    # ============================================================
    r = 34
    write_label(ws, r, "Finance income")
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        write_formula(ws, r, 2 + t, f"={C('a.finance_income_pct')}*{col}{IS_ROWS['revenue']}", NUM_BN)
        reg(f"is.finance_income_t{t}", SH_IS, col, r)
        reg(f"is.fin_inc_t{t}", SH_IS, col, r)  # back-compat
    IS_ROWS["finance_income"] = r
    IS_ROWS["fin_inc"] = r

    r = 35
    write_label(ws, r, "Finance costs (interest expense)")
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        write_formula(ws, r, 2 + t, f"=-{C(f'd.interest_t{t}')}", NUM_BN)
        reg(f"is.finance_costs_t{t}", SH_IS, col, r)
        reg(f"is.interest_t{t}", SH_IS, col, r)  # back-compat
    IS_ROWS["finance_costs"] = r
    IS_ROWS["interest"] = r

    r = 36
    write_label(ws, r, "Share of associates / JVs")
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        formula = f"={C('a.associates_y0')}*(1+{C('a.associates_growth')})^{t}"
        write_formula(ws, r, 2 + t, formula, NUM_BN)
        reg(f"is.associates_t{t}", SH_IS, col, r)
    IS_ROWS["associates"] = r

    r = 37
    write_label(ws, r, "Profit before tax", bold=True)
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        cells = [IS_ROWS[k] for k in ("ebit", "finance_income", "finance_costs", "associates")]
        formula = "=" + "+".join(f"{col}{x}" for x in cells)
        write_formula(ws, r, 2 + t, formula, NUM_BN, bold=True)
        reg(f"is.pretax_t{t}", SH_IS, col, r)
    IS_ROWS["pretax"] = r

    # ============================================================
    # SECTION 5 — TAX & NET INCOME (rows 38-40)
    # ============================================================
    r = 38
    write_label(ws, r, "Current tax expense")
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        write_formula(ws, r, 2 + t, f"=-{C(f't.expense_t{t}')}", NUM_BN)
        reg(f"is.current_tax_t{t}", SH_IS, col, r)
        reg(f"is.tax_t{t}", SH_IS, col, r)  # back-compat = current tax (deferred = 0 in Phase 1)
    IS_ROWS["current_tax"] = r
    IS_ROWS["tax"] = r

    r = 39
    write_label(ws, r, "Deferred tax (movement)")
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        write_formula(ws, r, 2 + t, "=0", NUM_BN)
        reg(f"is.deferred_tax_t{t}", SH_IS, col, r)
    IS_ROWS["deferred_tax"] = r

    r = 40
    write_label(ws, r, "NET INCOME", bold=True, banded=True)
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        cells = [IS_ROWS[k] for k in ("pretax", "current_tax", "deferred_tax")]
        formula = "=" + "+".join(f"{col}{x}" for x in cells)
        write_formula(ws, r, 2 + t, formula, NUM_BN, bold=True, banded=True)
        reg(f"is.ni_t{t}", SH_IS, col, r)
    IS_ROWS["ni"] = r

    # ============================================================
    # SECTION 6 — ATTRIBUTION & EPS (rows 42-45)
    # ============================================================
    r = 42
    write_label(ws, r, "Attributable to equity holders of the parent")
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        formula = f"={col}{IS_ROWS['ni']}*(1-{C('a.minority_share_ni')})"
        write_formula(ws, r, 2 + t, formula, NUM_BN)
        reg(f"is.attr_equity_t{t}", SH_IS, col, r)
    IS_ROWS["attr_equity"] = r

    r = 43
    write_label(ws, r, "Attributable to minority interest")
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        formula = f"={col}{IS_ROWS['ni']}*{C('a.minority_share_ni')}"
        write_formula(ws, r, 2 + t, formula, NUM_BN)
        reg(f"is.attr_minority_t{t}", SH_IS, col, r)
    IS_ROWS["attr_minority"] = r

    r = 45
    write_label(ws, r, "Basic EPS (SAR / share)")
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        # EPS = attributable NI (SAR bn) × 1000 / shares (millions) = SAR/share
        formula = f"={col}{IS_ROWS['attr_equity']}*1000/{C('a.shares')}"
        write_formula(ws, r, 2 + t, formula, '#,##0.00')
        reg(f"is.eps_t{t}", SH_IS, col, r)
    IS_ROWS["eps"] = r

    # ============================================================
    # SECTION 7 — MEMO (rows 47-51)
    # ============================================================
    section_header(ws, 47, "MEMO")

    for r, label, key, top_key in [
        (48, "EBITDA margin",        "ebitda_margin", "ebitda"),
        (49, "EBIT margin",          "ebit_margin",   "ebit"),
        (50, "Net margin",           "net_margin",    "ni"),
    ]:
        write_label(ws, r, label, indent=1)
        for t in range(1, 6):
            col = get_column_letter(2 + t)
            formula = f"={col}{IS_ROWS[top_key]}/{col}{IS_ROWS['revenue']}"
            write_formula(ws, r, 2 + t, formula, NUM_PCT)
        IS_ROWS[key] = r

    r = 51
    write_label(ws, r, "Effective tax rate", indent=1)
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        formula = f"=-({col}{IS_ROWS['current_tax']}+{col}{IS_ROWS['deferred_tax']})/{col}{IS_ROWS['pretax']}"
        write_formula(ws, r, 2 + t, formula, NUM_PCT)
    IS_ROWS["etr"] = r


# =============================================================================
# 15. EQUITY ROLLFORWARD
# =============================================================================

EQ_ROWS = {}

def build_equity(ws):
    ws.column_dimensions["A"].width = 32
    for c in range(2, 8):
        ws.column_dimensions[get_column_letter(c)].width = 13

    ws["A1"] = "EQUITY ROLLFORWARD"
    ws["A1"].font = section_font
    ws["A2"] = "Closing equity = opening + net income − dividends + OCI movements. Drives BS equity section."
    ws["A2"].font = sub_font

    year_header(ws, 4, 0, 5)
    r = 6

    # Share capital (flat)
    write_label(ws, r, "Share capital")
    write_formula(ws, r, 2, f"={C('a.share_capital')}", NUM_BN)
    for t in range(1, 6):
        prev = get_column_letter(2 + t - 1)
        write_formula(ws, r, 2 + t, f"={prev}{r}", NUM_BN)
    EQ_ROWS["share_cap"] = r; r += 1
    for t in range(0, 6):
        col = get_column_letter(2 + t)
        reg(f"eq.share_cap_t{t}", SH_EQ, col, EQ_ROWS["share_cap"])

    # Retained earnings rollforward
    write_label(ws, r, "Opening retained earnings")
    write_formula(ws, r, 2, '""', '@')
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        if t == 1:
            write_formula(ws, r, 2 + t, f"={C('a.re_y0')}", NUM_BN)
        else:
            prev_col = get_column_letter(2 + t - 1)
            write_formula(ws, r, 2 + t, f"={prev_col}{r + 3}", NUM_BN)  # prev closing RE
    EQ_ROWS["re_open"] = r; r += 1

    write_label(ws, r, "+ Net income for period")
    write_formula(ws, r, 2, '""', '@')
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        write_formula(ws, r, 2 + t, f"={C(f'is.ni_t{t}')}", NUM_BN)
    EQ_ROWS["ni"] = r; r += 1

    write_label(ws, r, "− Dividends paid")
    write_formula(ws, r, 2, '""', '@')
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        formula = f"=-{C('a.div_payout')}*{C(f'is.ni_t{t}')}"
        write_formula(ws, r, 2 + t, formula, NUM_BN)
        reg(f"eq.div_t{t}", SH_EQ, col, r)
    EQ_ROWS["div"] = r; r += 1

    write_label(ws, r, "Closing retained earnings", bold=True, banded=True)
    write_formula(ws, r, 2, f"={C('a.re_y0')}", NUM_BN, bold=True, banded=True)
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        formula = f"={col}{EQ_ROWS['re_open']}+{col}{EQ_ROWS['ni']}+{col}{EQ_ROWS['div']}"
        write_formula(ws, r, 2 + t, formula, NUM_BN, bold=True, banded=True)
    EQ_ROWS["re_close"] = r
    for t in range(0, 6):
        col = get_column_letter(2 + t)
        reg(f"eq.re_t{t}", SH_EQ, col, EQ_ROWS["re_close"])
    r += 2

    # OCI reserve (held flat in default)
    write_label(ws, r, "OCI reserve (held flat)")
    write_formula(ws, r, 2, f"={C('a.oci_reserve_y0')}", NUM_BN)
    for t in range(1, 6):
        prev = get_column_letter(2 + t - 1)
        write_formula(ws, r, 2 + t, f"={prev}{r}", NUM_BN)
    EQ_ROWS["oci"] = r
    for t in range(0, 6):
        col = get_column_letter(2 + t)
        reg(f"eq.oci_t{t}", SH_EQ, col, EQ_ROWS["oci"])
    r += 1

    # Minority interest (held flat)
    write_label(ws, r, "Minority interest (held flat)")
    write_formula(ws, r, 2, f"={C('a.minority_y0')}", NUM_BN)
    for t in range(1, 6):
        prev = get_column_letter(2 + t - 1)
        write_formula(ws, r, 2 + t, f"={prev}{r}", NUM_BN)
    EQ_ROWS["minority"] = r
    for t in range(0, 6):
        col = get_column_letter(2 + t)
        reg(f"eq.minority_t{t}", SH_EQ, col, EQ_ROWS["minority"])
    r += 2

    # Total equity
    write_label(ws, r, "TOTAL EQUITY", bold=True, banded=True)
    for t in range(0, 6):
        col = get_column_letter(2 + t)
        formula = f"={col}{EQ_ROWS['share_cap']}+{col}{EQ_ROWS['re_close']}+{col}{EQ_ROWS['oci']}+{col}{EQ_ROWS['minority']}"
        write_formula(ws, r, 2 + t, formula, NUM_BN, bold=True, banded=True)
        reg(f"eq.total_t{t}", SH_EQ, col, r)
    EQ_ROWS["total"] = r


# =============================================================================
# 16. BALANCE SHEET
# =============================================================================

BS_ROWS = {}

def build_bs(ws):
    ws.column_dimensions["A"].width = 34
    for c in range(2, 8):
        ws.column_dimensions[get_column_letter(c)].width = 13

    ws["A1"] = "BALANCE SHEET"
    ws["A1"].font = section_font
    ws["A2"] = "All figures in SAR billion. Y0 opening + Y1-Y5 forecast. Must balance (see Checks)."
    ws["A2"].font = sub_font

    year_header(ws, 4, 0, 5)

    r = 6
    # ---- ASSETS ----
    section_header(ws, r, "ASSETS"); r += 1

    # Cash (plug, from CFS)
    write_label(ws, r, "Cash & equivalents")
    write_formula(ws, r, 2, f"={C('a.cash_y0')}", NUM_BN)
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        prev = get_column_letter(2 + t - 1)
        write_formula(ws, r, 2 + t, f"={prev}{r}+CFS!{col}{CFS_PLACEHOLDER_NETCASH}", NUM_BN)
    BS_ROWS["cash"] = r
    for t in range(0, 6):
        col = get_column_letter(2 + t)
        reg(f"b.cash_t{t}", SH_BS, col, r)
    r += 1
    # Patch debt schedule placeholder for cash row
    global BS_PLACEHOLDER_CASH
    BS_PLACEHOLDER_CASH = r - 1  # because we already advanced r

    # AR
    write_label(ws, r, "Accounts receivable")
    write_formula(ws, r, 2, f"={C('a.ar_y0')}", NUM_BN)
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        write_formula(ws, r, 2 + t, f"={C(f'w.ar_t{t}')}", NUM_BN)
    BS_ROWS["ar"] = r; r += 1

    write_label(ws, r, "Inventory")
    write_formula(ws, r, 2, f"={C('a.inv_y0')}", NUM_BN)
    for t in range(1, 6):
        write_formula(ws, r, 2 + t, f"={C(f'w.inv_t{t}')}", NUM_BN)
    BS_ROWS["inv"] = r; r += 1

    write_label(ws, r, "Other current assets (flat)")
    write_formula(ws, r, 2, f"={C('a.other_ca_y0')}", NUM_BN)
    for t in range(1, 6):
        prev = get_column_letter(2 + t - 1)
        write_formula(ws, r, 2 + t, f"={prev}{r}", NUM_BN)
    BS_ROWS["other_ca"] = r; r += 1

    # Current assets total
    write_label(ws, r, "Total current assets", bold=True)
    for t in range(0, 6):
        col = get_column_letter(2 + t)
        formula = f"={col}{BS_ROWS['cash']}+{col}{BS_ROWS['ar']}+{col}{BS_ROWS['inv']}+{col}{BS_ROWS['other_ca']}"
        write_formula(ws, r, 2 + t, formula, NUM_BN, bold=True)
    BS_ROWS["ca_total"] = r; r += 2

    # Non-current assets
    write_label(ws, r, "PP&E")
    write_formula(ws, r, 2, f"={C('a.ppe_y0')}", NUM_BN)
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        write_formula(ws, r, 2 + t, f"={C(f'p.closing_t{t}')}", NUM_BN)
    BS_ROWS["ppe"] = r; r += 1

    write_label(ws, r, "Goodwill (flat)")
    write_formula(ws, r, 2, f"={C('a.goodwill_y0')}", NUM_BN)
    for t in range(1, 6):
        prev = get_column_letter(2 + t - 1)
        write_formula(ws, r, 2 + t, f"={prev}{r}", NUM_BN)
    BS_ROWS["goodwill"] = r; r += 1

    write_label(ws, r, "Intangibles (less amortisation)")
    write_formula(ws, r, 2, f"={C('a.intangibles_y0')}", NUM_BN)
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        prev = get_column_letter(2 + t - 1)
        # IS amortisation is recorded negative; adding it = subtracting magnitude
        write_formula(ws, r, 2 + t, f"={prev}{r}+{C(f'is.amortisation_t{t}')}", NUM_BN)
    BS_ROWS["intangibles"] = r; r += 1

    write_label(ws, r, "Investments / associates (flat)")
    write_formula(ws, r, 2, f"={C('a.investments_y0')}", NUM_BN)
    for t in range(1, 6):
        prev = get_column_letter(2 + t - 1)
        write_formula(ws, r, 2 + t, f"={prev}{r}", NUM_BN)
    BS_ROWS["investments"] = r; r += 1

    write_label(ws, r, "Deferred tax assets (flat)")
    write_formula(ws, r, 2, f"={C('a.dta_y0')}", NUM_BN)
    for t in range(1, 6):
        prev = get_column_letter(2 + t - 1)
        write_formula(ws, r, 2 + t, f"={prev}{r}", NUM_BN)
    BS_ROWS["dta"] = r; r += 1

    write_label(ws, r, "Total non-current assets", bold=True)
    for t in range(0, 6):
        col = get_column_letter(2 + t)
        cells = [BS_ROWS["ppe"], BS_ROWS["goodwill"], BS_ROWS["intangibles"], BS_ROWS["investments"], BS_ROWS["dta"]]
        formula = "=" + "+".join(f"{col}{x}" for x in cells)
        write_formula(ws, r, 2 + t, formula, NUM_BN, bold=True)
    BS_ROWS["nca_total"] = r; r += 2

    # Total assets
    write_label(ws, r, "TOTAL ASSETS", bold=True, banded=True)
    for t in range(0, 6):
        col = get_column_letter(2 + t)
        formula = f"={col}{BS_ROWS['ca_total']}+{col}{BS_ROWS['nca_total']}"
        write_formula(ws, r, 2 + t, formula, NUM_BN, bold=True, banded=True)
        reg(f"b.total_assets_t{t}", SH_BS, col, r)
    BS_ROWS["total_assets"] = r; r += 2

    # ---- LIABILITIES ----
    section_header(ws, r, "LIABILITIES"); r += 1

    write_label(ws, r, "Accounts payable")
    write_formula(ws, r, 2, f"={C('a.ap_y0')}", NUM_BN)
    for t in range(1, 6):
        write_formula(ws, r, 2 + t, f"={C(f'w.ap_t{t}')}", NUM_BN)
    BS_ROWS["ap"] = r; r += 1

    write_label(ws, r, "Other current liabilities (flat)")
    write_formula(ws, r, 2, f"={C('a.other_cl_y0')}", NUM_BN)
    for t in range(1, 6):
        prev = get_column_letter(2 + t - 1)
        write_formula(ws, r, 2 + t, f"={prev}{r}", NUM_BN)
    BS_ROWS["other_cl"] = r; r += 1

    write_label(ws, r, "Short-term debt")
    for t in range(0, 6):
        col = get_column_letter(2 + t)
        write_formula(ws, r, 2 + t, f"={SH_DEBT}!{col}{DEBT_ROWS['st']}", NUM_BN)
    BS_ROWS["st_debt"] = r; r += 1

    write_label(ws, r, "Total current liabilities", bold=True)
    for t in range(0, 6):
        col = get_column_letter(2 + t)
        formula = f"={col}{BS_ROWS['ap']}+{col}{BS_ROWS['other_cl']}+{col}{BS_ROWS['st_debt']}"
        write_formula(ws, r, 2 + t, formula, NUM_BN, bold=True)
    BS_ROWS["cl_total"] = r; r += 2

    write_label(ws, r, "Long-term debt")
    for t in range(0, 6):
        col = get_column_letter(2 + t)
        write_formula(ws, r, 2 + t, f"={SH_DEBT}!{col}{DEBT_ROWS['lt']}", NUM_BN)
    BS_ROWS["lt_debt"] = r; r += 1

    write_label(ws, r, "Lease liabilities")
    for t in range(0, 6):
        col = get_column_letter(2 + t)
        write_formula(ws, r, 2 + t, f"={SH_DEBT}!{col}{DEBT_ROWS['lease']}", NUM_BN)
    BS_ROWS["lease"] = r; r += 1

    write_label(ws, r, "Provisions (flat)")
    write_formula(ws, r, 2, f"={C('a.provisions_y0')}", NUM_BN)
    for t in range(1, 6):
        prev = get_column_letter(2 + t - 1)
        write_formula(ws, r, 2 + t, f"={prev}{r}", NUM_BN)
    BS_ROWS["provisions"] = r; r += 1

    write_label(ws, r, "Deferred tax liabilities (flat)")
    write_formula(ws, r, 2, f"={C('a.dtl_y0')}", NUM_BN)
    for t in range(1, 6):
        prev = get_column_letter(2 + t - 1)
        write_formula(ws, r, 2 + t, f"={prev}{r}", NUM_BN)
    BS_ROWS["dtl"] = r; r += 1

    write_label(ws, r, "Other non-current liabilities (flat)")
    write_formula(ws, r, 2, f"={C('a.other_ncl_y0')}", NUM_BN)
    for t in range(1, 6):
        prev = get_column_letter(2 + t - 1)
        write_formula(ws, r, 2 + t, f"={prev}{r}", NUM_BN)
    BS_ROWS["other_ncl"] = r; r += 1

    write_label(ws, r, "Total non-current liabilities", bold=True)
    for t in range(0, 6):
        col = get_column_letter(2 + t)
        cells = [BS_ROWS["lt_debt"], BS_ROWS["lease"], BS_ROWS["provisions"], BS_ROWS["dtl"], BS_ROWS["other_ncl"]]
        formula = "=" + "+".join(f"{col}{x}" for x in cells)
        write_formula(ws, r, 2 + t, formula, NUM_BN, bold=True)
    BS_ROWS["ncl_total"] = r; r += 2

    write_label(ws, r, "TOTAL LIABILITIES", bold=True, banded=True)
    for t in range(0, 6):
        col = get_column_letter(2 + t)
        formula = f"={col}{BS_ROWS['cl_total']}+{col}{BS_ROWS['ncl_total']}"
        write_formula(ws, r, 2 + t, formula, NUM_BN, bold=True, banded=True)
        reg(f"b.total_liab_t{t}", SH_BS, col, r)
    BS_ROWS["total_liab"] = r; r += 2

    # ---- EQUITY ----
    section_header(ws, r, "EQUITY"); r += 1

    write_label(ws, r, "Share capital")
    for t in range(0, 6):
        col = get_column_letter(2 + t)
        write_formula(ws, r, 2 + t, f"={C(f'eq.share_cap_t{t}')}", NUM_BN)
    r += 1
    write_label(ws, r, "Retained earnings")
    for t in range(0, 6):
        col = get_column_letter(2 + t)
        write_formula(ws, r, 2 + t, f"={C(f'eq.re_t{t}')}", NUM_BN)
    r += 1
    write_label(ws, r, "OCI reserve")
    for t in range(0, 6):
        col = get_column_letter(2 + t)
        write_formula(ws, r, 2 + t, f"={C(f'eq.oci_t{t}')}", NUM_BN)
    r += 1
    write_label(ws, r, "Minority interest")
    for t in range(0, 6):
        col = get_column_letter(2 + t)
        write_formula(ws, r, 2 + t, f"={C(f'eq.minority_t{t}')}", NUM_BN)
    r += 1
    write_label(ws, r, "TOTAL EQUITY", bold=True, banded=True)
    for t in range(0, 6):
        col = get_column_letter(2 + t)
        write_formula(ws, r, 2 + t, f"={C(f'eq.total_t{t}')}", NUM_BN, bold=True, banded=True)
        reg(f"b.total_equity_t{t}", SH_BS, col, r)
    BS_ROWS["total_equity"] = r; r += 2

    # Check row: A - L - E
    write_label(ws, r, "Balance check (A − L − E)", bold=True)
    for t in range(0, 6):
        col = get_column_letter(2 + t)
        formula = f"={col}{BS_ROWS['total_assets']}-{col}{BS_ROWS['total_liab']}-{col}{BS_ROWS['total_equity']}"
        write_formula(ws, r, 2 + t, formula, NUM_BN, bold=True)
    BS_ROWS["check"] = r


CFS_PLACEHOLDER_NETCASH = 0


# =============================================================================
# 17. CASH FLOW STATEMENT
# =============================================================================

CFS_ROWS = {}

def build_cfs(ws):
    ws.column_dimensions["A"].width = 34
    for c in range(2, 8):
        ws.column_dimensions[get_column_letter(c)].width = 13

    ws["A1"] = "CASH FLOW STATEMENT (indirect method)"
    ws["A1"].font = section_font
    ws["A2"] = "All figures in SAR billion. Net change in cash flows to BS."
    ws["A2"].font = sub_font

    ws.cell(row=4, column=1).fill = header_fill
    for t in range(1, 6):
        c = ws.cell(row=4, column=2 + t, value=f"Y{t}")
        c.font = header_font; c.fill = header_fill; c.alignment = right

    r = 6
    section_header(ws, r, "OPERATING"); r += 1

    write_label(ws, r, "Net income")
    for t in range(1, 6):
        write_formula(ws, r, 2 + t, f"={C(f'is.ni_t{t}')}", NUM_BN)
    CFS_ROWS["ni"] = r; r += 1

    write_label(ws, r, "+ Depreciation (non-cash addback)")
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        # depreciation on IS is negative; we add back the absolute value
        write_formula(ws, r, 2 + t, f"=-{C(f'is.depreciation_t{t}')}", NUM_BN)
    CFS_ROWS["da"] = r; r += 1

    write_label(ws, r, "+ Amortisation (non-cash addback)")
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        write_formula(ws, r, 2 + t, f"=-{C(f'is.amortisation_t{t}')}", NUM_BN)
    CFS_ROWS["amort"] = r; r += 1

    write_label(ws, r, "+/− Δ working capital")
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        write_formula(ws, r, 2 + t, f"={C(f'w.dwc_t{t}')}", NUM_BN)
    CFS_ROWS["dwc"] = r; r += 1

    write_label(ws, r, "Cash from operations (CFO)", bold=True, banded=True)
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        formula = (f"={col}{CFS_ROWS['ni']}+{col}{CFS_ROWS['da']}"
                   f"+{col}{CFS_ROWS['amort']}+{col}{CFS_ROWS['dwc']}")
        write_formula(ws, r, 2 + t, formula, NUM_BN, bold=True, banded=True)
        reg(f"c.cfo_t{t}", SH_CFS, col, r)
    CFS_ROWS["cfo"] = r; r += 2

    section_header(ws, r, "INVESTING"); r += 1

    write_label(ws, r, "− Capex")
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        write_formula(ws, r, 2 + t, f"=-{C(f'p.capex_t{t}')}", NUM_BN)
    CFS_ROWS["capex"] = r; r += 1

    write_label(ws, r, "Other investing (flat / nil)")
    for t in range(1, 6):
        write_formula(ws, r, 2 + t, "=0", NUM_BN)
    CFS_ROWS["other_inv"] = r; r += 1

    write_label(ws, r, "Cash from investing (CFI)", bold=True, banded=True)
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        formula = f"={col}{CFS_ROWS['capex']}+{col}{CFS_ROWS['other_inv']}"
        write_formula(ws, r, 2 + t, formula, NUM_BN, bold=True, banded=True)
        reg(f"c.cfi_t{t}", SH_CFS, col, r)
    CFS_ROWS["cfi"] = r; r += 2

    section_header(ws, r, "FINANCING"); r += 1

    write_label(ws, r, "+/− Net debt issuance (flat assumption)")
    for t in range(1, 6):
        write_formula(ws, r, 2 + t, "=0", NUM_BN)  # debt held flat
    CFS_ROWS["debt_chg"] = r; r += 1

    write_label(ws, r, "− Dividends paid")
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        write_formula(ws, r, 2 + t, f"={C(f'eq.div_t{t}')}", NUM_BN)
    CFS_ROWS["div"] = r; r += 1

    write_label(ws, r, "Cash from financing (CFF)", bold=True, banded=True)
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        formula = f"={col}{CFS_ROWS['debt_chg']}+{col}{CFS_ROWS['div']}"
        write_formula(ws, r, 2 + t, formula, NUM_BN, bold=True, banded=True)
        reg(f"c.cff_t{t}", SH_CFS, col, r)
    CFS_ROWS["cff"] = r; r += 2

    write_label(ws, r, "Net change in cash", bold=True, banded=True)
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        formula = f"={col}{CFS_ROWS['cfo']}+{col}{CFS_ROWS['cfi']}+{col}{CFS_ROWS['cff']}"
        write_formula(ws, r, 2 + t, formula, NUM_BN, bold=True, banded=True)
        reg(f"c.netcash_t{t}", SH_CFS, col, r)
    CFS_ROWS["netcash"] = r

    # Patch BS placeholder so cash row picks up the netcash
    global CFS_PLACEHOLDER_NETCASH
    CFS_PLACEHOLDER_NETCASH = r

    r += 1
    write_label(ws, r, "Opening cash")
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        if t == 1:
            write_formula(ws, r, 2 + t, f"={C('a.cash_y0')}", NUM_BN)
        else:
            prev = get_column_letter(2 + t - 1)
            write_formula(ws, r, 2 + t, f"={prev}{r + 1}", NUM_BN)
    CFS_ROWS["open_cash"] = r; r += 1

    write_label(ws, r, "Closing cash", bold=True, banded=True)
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        formula = f"={col}{CFS_ROWS['open_cash']}+{col}{CFS_ROWS['netcash']}"
        write_formula(ws, r, 2 + t, formula, NUM_BN, bold=True, banded=True)
    CFS_ROWS["close_cash"] = r


# =============================================================================
# 18. BUDGET (Y1 monthly)
# =============================================================================

def build_budget(ws):
    ws.column_dimensions["A"].width = 30
    for c in range(2, 16):
        ws.column_dimensions[get_column_letter(c)].width = 11

    ws["A1"] = "Y1 MONTHLY BUDGET"
    ws["A1"].font = section_font
    ws["A2"] = "Y1 P&L line items, distributed by month per the seasonality weights below. Variance columns are placeholders for actuals."
    ws["A2"].font = sub_font

    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

    # Header
    ws.cell(row=4, column=1).fill = header_fill
    for i, m in enumerate(months):
        c = ws.cell(row=4, column=2 + i, value=m)
        c.font = header_font; c.fill = header_fill; c.alignment = right
    c = ws.cell(row=4, column=14, value="Y1 Total")
    c.font = header_font; c.fill = header_fill; c.alignment = right

    # Weights row
    r = 5
    write_label(ws, r, "Seasonality weight")
    for i, w in enumerate(D["month_weights"]):
        write_input(ws, r, 2 + i, w, NUM_PCT)
    write_formula(ws, r, 14, f"=SUM(B{r}:M{r})", NUM_PCT, bold=True)
    weights_row = r; r += 2

    # P&L lines (using IS Y1 references)
    section_header(ws, r, "P&L lines (SAR bn)"); r += 1

    line_defs = [
        ("Total revenue",             f"={C('is.total_revenue_t1')}"),
        ("  Service revenue",         f"={C('is.service_revenue_t1')}"),
        ("  Equipment revenue",       f"={C('is.equipment_revenue_t1')}"),
        ("Interconnect & roaming",    f"={C('is.interconnect_t1')}"),
        ("Content / OTT",             f"={C('is.content_t1')}"),
        ("Equipment COGS",            f"={C('is.equipment_cogs_t1')}"),
        ("Network energy + lease",    f"={C('is.energy_lease_t1')}"),
        ("Other network opex",        f"={C('is.other_network_t1')}"),
        ("Employee costs",            f"={C('is.employee_t1')}"),
        ("Customer acquisition",      f"={C('is.cust_acq_t1')}"),
        ("Other commercial",          f"={C('is.commercial_other_t1')}"),
        ("G&A",                       f"={C('is.ga_t1')}"),
        ("FX losses",                 f"={C('is.fx_t1')}"),
        ("EBITDA",                    f"={C('is.ebitda_t1')}"),
        ("Depreciation",              f"={C('is.depreciation_t1')}"),
        ("Amortisation",              f"={C('is.amortisation_t1')}"),
        ("EBIT",                      f"={C('is.ebit_t1')}"),
        ("Net income",                f"={C('is.ni_t1')}"),
        ("  Attributable to parent",  f"={C('is.attr_equity_t1')}"),
        ("  Minority interest",       f"={C('is.attr_minority_t1')}"),
    ]
    for label, y1_formula in line_defs:
        bold = label in ("EBITDA", "EBIT", "Net income")
        write_label(ws, r, label, bold=bold, banded=bold)
        # Per-month = weight × annual
        for i in range(12):
            col = get_column_letter(2 + i)
            formula = f"={col}{weights_row}*({y1_formula[1:]})"
            write_formula(ws, r, 2 + i, formula, NUM_BN, bold=bold, banded=bold)
        # Total = SUM
        write_formula(ws, r, 14, f"=SUM(B{r}:M{r})", NUM_BN, bold=True, banded=bold)
        r += 1

    r += 1
    section_header(ws, r, "Quarterly summary"); r += 1
    write_label(ws, r, "Service revenue (quarterly)")
    # Columns N..Q for quarterly. Actually simpler: print Q1..Q4 in rows.
    ws.cell(row=r, column=2, value="Q1"); ws.cell(row=r, column=2).font = sub_font
    ws.cell(row=r, column=3, value="Q2"); ws.cell(row=r, column=3).font = sub_font
    ws.cell(row=r, column=4, value="Q3"); ws.cell(row=r, column=4).font = sub_font
    ws.cell(row=r, column=5, value="Q4"); ws.cell(row=r, column=5).font = sub_font
    r += 1
    write_label(ws, r, "Quarterly service revenue")
    # First service revenue row is r-something; we tracked first line above
    sr_row = weights_row + 2  # first line under "P&L lines" section header
    for q, cols in enumerate([(2,3,4),(5,6,7),(8,9,10),(11,12,13)]):
        col = get_column_letter(2 + q)
        terms = "+".join(f"{get_column_letter(c)}{sr_row}" for c in cols)
        write_formula(ws, r, 2 + q, f"={terms}", NUM_BN, bold=True, banded=True)


# =============================================================================
# 19. SCENARIOS
# =============================================================================

def build_scenarios(ws):
    ws.column_dimensions["A"].width = 30
    for c in range(2, 6):
        ws.column_dimensions[get_column_letter(c)].width = 14

    ws["A1"] = "SCENARIO MANAGER"
    ws["A1"].font = section_font
    ws["A2"] = "Reference table for downside / base / upside parameter sets. Active scenario set on Assumptions sheet."
    ws["A2"].font = sub_font

    # Header
    ws.cell(row=4, column=1).fill = header_fill
    for i, name in enumerate(["1: Downside", "2: Base", "3: Upside"]):
        c = ws.cell(row=4, column=2 + i, value=name)
        c.font = header_font; c.fill = header_fill; c.alignment = right

    r = 6
    write_label(ws, r, "Revenue growth")
    for i, v in enumerate(D["scen_growth"]):
        write_input(ws, r, 2 + i, v, NUM_PCT)
    r += 1
    write_label(ws, r, "EBITDA margin (target)")
    for i, v in enumerate(D["scen_margin"]):
        write_input(ws, r, 2 + i, v, NUM_PCT)
    r += 1
    write_label(ws, r, "Capex intensity")
    for i, v in enumerate(D["scen_capex"]):
        write_input(ws, r, 2 + i, v, NUM_PCT)
    r += 2

    write_label(ws, r, "Active scenario (1/2/3)", bold=True)
    write_formula(ws, r, 2, f"={C('a.scen_default')}", NUM_INT, bold=True, banded=True)
    r += 2

    section_header(ws, r, "Active values (use INDEX in Assumptions to wire these)"); r += 1
    write_label(ws, r, "Note")
    note = ws.cell(row=r, column=2, value=("Manually overwrite Assumptions!B6/B7/B9 with these "
                                          "values, or build INDEX formulas. Kept manual for clarity."))
    note.font = sub_font; note.alignment = left
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)


# =============================================================================
# 20. SENSITIVITY
# =============================================================================

def build_sensitivity(ws):
    ws.column_dimensions["A"].width = 30
    for c in range(2, 10):
        ws.column_dimensions[get_column_letter(c)].width = 12

    ws["A1"] = "SENSITIVITY TABLES"
    ws["A1"].font = section_font
    ws["A2"] = "Two-way tables on the key valuation levers. Outputs computed via TABLE-style formulas — change axis values to flex."
    ws["A2"].font = sub_font

    # ---- Table 1: WACC × g => Per-share value ----
    r = 4
    section_header(ws, r, "Per-share value (SAR) — WACC × terminal growth"); r += 1

    waccs = [0.07, 0.08, 0.09, 0.10, 0.11]
    gs    = [0.01, 0.02, 0.03, 0.04]

    # Header row (WACCs)
    write_label(ws, r, "WACC →")
    for i, w in enumerate(waccs):
        write_input(ws, r, 2 + i, w, NUM_PCT)
    r += 1
    # Body
    for j, g in enumerate(gs):
        write_input(ws, r, 1, g, NUM_PCT)
        # We compute approximate per-share by replicating the DCF math in-formula.
        for i, w in enumerate(waccs):
            # Use IS revenue Y5, EBITDA margin from a.ebitda_target, derive simple FCFF
            # For simplicity reference Valuation EV computation pattern but with these w, g
            # Per-share = (PV(explicit FCFF using w) + PV(TV using w,g) - net_debt) * 1000 / shares
            # We construct the formula inline. Since FCFF depends on a.wacc, we cannot
            # use Valuation sheet directly. Replicate inline:
            terms = []
            for t in range(1, 6):
                ebit = f"{C(f'is.ebit_t{t}')}"
                da = f"(-{C(f'is.da_t{t}')})"
                capex = f"{C(f'p.capex_t{t}')}"
                dwc = f"{C(f'w.dwc_t{t}')}"
                fcff = f"({ebit}*(1-{C('a.tax')})+{da}-{capex}+{dwc})"
                terms.append(f"({fcff}/(1+{get_column_letter(2 + i)}{r - 1 - j})^{t})")
            sum_fcff = "+".join(terms)
            tv = (f"(({C('is.ebit_t5')}*(1-{C('a.tax')})+(-{C('is.da_t5')})-{C('p.capex_t5')}+{C('w.dwc_t5')})"
                  f"*(1+A{r})/({get_column_letter(2 + i)}{r - 1 - j}-A{r}))")
            pv_tv = f"({tv}/(1+{get_column_letter(2 + i)}{r - 1 - j})^5)"
            ev_expr = f"({sum_fcff}+{pv_tv})"
            equity_expr = f"({ev_expr}-{C('d.net_y0')})"
            ps_expr = f"={equity_expr}*1000/{C('a.shares')}"
            write_formula(ws, r, 2 + i, ps_expr, '#,##0.0')
        r += 1

    r += 2
    section_header(ws, r, "Enterprise value (SAR bn) — Y1 growth × EBITDA target margin"); r += 1
    growths = [0.02, 0.04, 0.06, 0.08, 0.10]
    margins = [0.30, 0.33, 0.36, 0.39, 0.42]
    write_label(ws, r, "Growth →")
    for i, gv in enumerate(growths):
        write_input(ws, r, 2 + i, gv, NUM_PCT)
    r += 1
    # Body: highly simplified — EV ≈ FCFF_Y5 / (wacc - g_terminal) where FCFF_Y5 derived
    # from simple formula: rev0*(1+growth)^5 * margin * (1-tax) - rev0*(1+growth)^5 * capex_pct
    for j, m in enumerate(margins):
        write_input(ws, r, 1, m, NUM_PCT)
        for i, gv in enumerate(growths):
            rev5 = f"({C('a.rev0')}*(1+{get_column_letter(2 + i)}{r - 1 - j})^5)"
            fcff5 = f"({rev5}*({m}-{C('a.capex_pct')})*(1-{C('a.tax')}))"
            tv = f"({fcff5}*(1+{C('a.g')})/({C('a.wacc')}-{C('a.g')}))"
            pv_tv = f"({tv}/(1+{C('a.wacc')})^5)"
            # Quick PV explicit: 5x FCFF5 * discount (rough)
            sum_pv = f"({fcff5}*5/(1+{C('a.wacc')})^3)"  # simplification
            write_formula(ws, r, 2 + i, f"={sum_pv}+{pv_tv}", NUM_BN)
        r += 1


# =============================================================================
# 21. VALUATION
# =============================================================================

VAL_ROWS = {}

def build_valuation(ws):
    ws.column_dimensions["A"].width = 32
    for c in range(2, 8):
        ws.column_dimensions[get_column_letter(c)].width = 13

    ws["A1"] = "DCF VALUATION"
    ws["A1"].font = section_font
    ws["A2"] = "Free cash flow to firm (FCFF), discounted at WACC. Terminal value via Gordon growth."
    ws["A2"].font = sub_font

    ws.cell(row=4, column=1).fill = header_fill
    for t in range(1, 6):
        c = ws.cell(row=4, column=2 + t, value=f"Y{t}")
        c.font = header_font; c.fill = header_fill; c.alignment = right

    r = 6
    write_label(ws, r, "EBIT (from IS)")
    for t in range(1, 6):
        write_formula(ws, r, 2 + t, f"={C(f'is.ebit_t{t}')}", NUM_BN)
    VAL_ROWS["ebit"] = r; r += 1

    write_label(ws, r, "× (1 − tax)")
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        write_formula(ws, r, 2 + t, f"={col}{VAL_ROWS['ebit']}*(1-{C('a.tax')})", NUM_BN)
    VAL_ROWS["nopat"] = r; r += 1

    write_label(ws, r, "+ Depreciation (non-cash)")
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        write_formula(ws, r, 2 + t, f"=-{C(f'is.depreciation_t{t}')}", NUM_BN)
    VAL_ROWS["da"] = r; r += 1

    write_label(ws, r, "+ Amortisation (non-cash)")
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        write_formula(ws, r, 2 + t, f"=-{C(f'is.amortisation_t{t}')}", NUM_BN)
    VAL_ROWS["amort"] = r; r += 1

    write_label(ws, r, "− Capex")
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        write_formula(ws, r, 2 + t, f"=-{C(f'p.capex_t{t}')}", NUM_BN)
    VAL_ROWS["capex"] = r; r += 1

    write_label(ws, r, "+/− Δ working capital")
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        write_formula(ws, r, 2 + t, f"={C(f'w.dwc_t{t}')}", NUM_BN)
    VAL_ROWS["dwc"] = r; r += 1

    write_label(ws, r, "FCFF", bold=True, banded=True)
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        formula = (f"={col}{VAL_ROWS['nopat']}+{col}{VAL_ROWS['da']}"
                   f"+{col}{VAL_ROWS['amort']}+{col}{VAL_ROWS['capex']}+{col}{VAL_ROWS['dwc']}")
        write_formula(ws, r, 2 + t, formula, NUM_BN, bold=True, banded=True)
    VAL_ROWS["fcff"] = r; r += 2

    write_label(ws, r, "Discount factor (1/(1+WACC)^t)")
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        write_formula(ws, r, 2 + t, f"=1/(1+{C('a.wacc')})^{t}", '0.0000')
    VAL_ROWS["df"] = r; r += 1

    write_label(ws, r, "PV of FCFF", bold=True)
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        formula = f"={col}{VAL_ROWS['fcff']}*{col}{VAL_ROWS['df']}"
        write_formula(ws, r, 2 + t, formula, NUM_BN, bold=True)
    VAL_ROWS["pv_fcff"] = r; r += 2

    # Summary block (col A label, col B value)
    write_label(ws, r, "Sum PV (explicit Y1-Y5)", bold=True)
    write_formula(ws, r, 2, f"=SUM(C{VAL_ROWS['pv_fcff']}:G{VAL_ROWS['pv_fcff']})", NUM_BN, bold=True)
    VAL_ROWS["sum_pv"] = r; r += 1

    write_label(ws, r, "Terminal value (Gordon)")
    write_formula(ws, r, 2, f"=G{VAL_ROWS['fcff']}*(1+{C('a.g')})/({C('a.wacc')}-{C('a.g')})", NUM_BN)
    VAL_ROWS["tv"] = r; r += 1

    write_label(ws, r, "PV of terminal value", bold=True)
    write_formula(ws, r, 2, f"=B{VAL_ROWS['tv']}*G{VAL_ROWS['df']}", NUM_BN, bold=True)
    VAL_ROWS["pv_tv"] = r; r += 2

    write_label(ws, r, "ENTERPRISE VALUE", bold=True, banded=True)
    write_formula(ws, r, 2, f"=B{VAL_ROWS['sum_pv']}+B{VAL_ROWS['pv_tv']}", NUM_BN, bold=True, banded=True)
    VAL_ROWS["ev"] = r
    reg("v.ev", SH_VAL, "B", r); r += 1

    write_label(ws, r, "− Net debt (Y0)")
    write_formula(ws, r, 2, f"=-{C('d.net_y0')}", NUM_BN)
    r += 1

    write_label(ws, r, "EQUITY VALUE", bold=True, banded=True)
    write_formula(ws, r, 2, f"=B{VAL_ROWS['ev']}+B{r-1}", NUM_BN, bold=True, banded=True)
    VAL_ROWS["equity"] = r; r += 1

    write_label(ws, r, "Per-share value (SAR)", bold=True, banded=True)
    write_formula(ws, r, 2, f"=B{VAL_ROWS['equity']}*1000/{C('a.shares')}", '#,##0.00', bold=True, banded=True)
    VAL_ROWS["per_share"] = r; r += 2

    section_header(ws, r, "Cross-checks"); r += 1
    write_label(ws, r, "Implied FY+1 EV/EBITDA")
    write_formula(ws, r, 2, f"=B{VAL_ROWS['ev']}/{C('is.ebitda_t1')}", NUM_X)
    r += 1
    write_label(ws, r, "Implied FY+1 EV/EBIT")
    write_formula(ws, r, 2, f"=B{VAL_ROWS['ev']}/{C('is.ebit_t1')}", NUM_X)
    r += 1
    write_label(ws, r, "Implied FY+1 P/E")
    write_formula(ws, r, 2, f"=B{VAL_ROWS['equity']}/{C('is.ni_t1')}", NUM_X)


# =============================================================================
# 22. CHECKS
# =============================================================================

def build_checks(ws):
    ws.column_dimensions["A"].width = 36
    for c in range(2, 10):
        ws.column_dimensions[get_column_letter(c)].width = 13

    ws["A1"] = "MODEL INTEGRITY CHECKS"
    ws["A1"].font = section_font
    ws["A2"] = "All checks should evaluate to PASS. Tolerance is 0.01 (SAR 10 million)."
    ws["A2"].font = sub_font

    year_header(ws, 4, 0, 5)

    r = 6
    # 1. BS balance check
    write_label(ws, r, "1. Balance sheet balances (A − L − E)")
    for t in range(0, 6):
        col = get_column_letter(2 + t)
        check = f"=IF(ABS({SH_BS}!{col}{BS_ROWS['check']})<0.01,\"PASS\",\"FAIL: \"&TEXT({SH_BS}!{col}{BS_ROWS['check']},\"+0.00;-0.00\"))"
        c = ws.cell(row=r, column=2 + t, value=check)
        c.alignment = right
        c.font = formula_font
    r += 1

    # 2. Cash on BS = closing cash on CFS rollforward
    write_label(ws, r, "2. BS cash = CFS rollforward")
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        check = (f"=IF(ABS({SH_BS}!{col}{BS_ROWS['cash']}-{SH_CFS}!{col}{CFS_ROWS['close_cash']})<0.01,"
                 f"\"PASS\",\"FAIL\")")
        c = ws.cell(row=r, column=2 + t, value=check)
        c.alignment = right
        c.font = formula_font
    r += 1

    # 3. Equity rollforward = BS equity
    write_label(ws, r, "3. Equity rollforward = BS equity")
    for t in range(0, 6):
        col = get_column_letter(2 + t)
        check = (f"=IF(ABS({SH_EQ}!{col}{EQ_ROWS['total']}-{SH_BS}!{col}{BS_ROWS['total_equity']})<0.01,"
                 f"\"PASS\",\"FAIL\")")
        c = ws.cell(row=r, column=2 + t, value=check)
        c.alignment = right
        c.font = formula_font
    r += 1

    # 4. Cash positivity
    write_label(ws, r, "4. Cash balance ≥ 0")
    for t in range(0, 6):
        col = get_column_letter(2 + t)
        check = f"=IF({SH_BS}!{col}{BS_ROWS['cash']}>=0,\"PASS\",\"FAIL (negative)\")"
        c = ws.cell(row=r, column=2 + t, value=check)
        c.alignment = right
        c.font = formula_font
    r += 1

    # 5. Net income > 0 in all years (just a sanity)
    write_label(ws, r, "5. Net income positive")
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        check = f"=IF({C(f'is.ni_t{t}')}>0,\"PASS\",\"WARNING\")"
        c = ws.cell(row=r, column=2 + t, value=check)
        c.alignment = right
        c.font = formula_font
    r += 1

    # Apply conditional formatting via openpyxl's CellIsRule
    from openpyxl.formatting.rule import CellIsRule, FormulaRule
    pass_dxf = pass_fill
    fail_dxf = fail_fill
    # Range covering all check cells: B6:G10
    rng = f"B6:G{r-1}"
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal",
                                                  formula=['"PASS"'],
                                                  fill=pass_fill,
                                                  font=pass_font))
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'AND(B6<>"PASS",B6<>"")'],
                                                   fill=fail_fill,
                                                   font=fail_font))


# =============================================================================
# 23. MAIN
# =============================================================================

def main():
    wb = Workbook()

    cover = wb.active
    cover.title = SH_COVER
    build_cover(cover)

    # Pre-register IS rows so downstream sheets (PPE, WC, Tax, Segments) can
    # reference them before the IS itself is constructed. build_is_v2 verifies
    # that actual rows match — if not, the script raises and refuses to write.
    PRE_IS_LAYOUT = {
        # Revenue block (rows 5-14)
        "seg_postpaid":      6,
        "seg_prepaid":       7,
        "seg_fixed":         8,
        "seg_b2b":           9,
        "seg_ict":          10,
        "seg_wholesale":    11,
        "service_revenue":  12,
        "equipment_revenue":13,
        "revenue":          14,   # = TOTAL revenue (service + equipment)
        "total_revenue":    14,
        # Cost block (rows 16-27)
        "interconnect":     17,
        "content":          18,
        "equipment_cogs":   19,
        "energy_lease":     20,
        "other_network":    21,
        "employee":         22,
        "cust_acq":         23,
        "commercial_other": 24,
        "ga":               25,
        "fx":               26,
        "total_opex":       27,
        # EBITDA & below
        "ebitda":           28,
        "depreciation":     30,
        "da":               30,   # back-compat alias = depreciation
        "amortisation":     31,
        "ebit":             32,
        "finance_income":   34,
        "fin_inc":          34,   # back-compat
        "finance_costs":    35,
        "interest":         35,   # back-compat
        "associates":       36,
        "pretax":           37,
        "current_tax":      38,
        "tax":              38,   # back-compat = current tax (deferred = 0)
        "deferred_tax":     39,
        "ni":               40,
        "attr_equity":      42,
        "attr_minority":    43,
        "eps":              45,
        "ebitda_margin":    48,
        "ebit_margin":      49,
        "net_margin":       50,
        "etr":              51,
    }
    for key, row in PRE_IS_LAYOUT.items():
        for t in range(1, 6):
            col = get_column_letter(2 + t)
            reg(f"is.{key}_t{t}", SH_IS, col, row)

    # Build order matters: shared registries get populated as we go.
    build_assumptions(wb.create_sheet(SH_ASSUM))
    build_drivers(wb.create_sheet(SH_DRIV))

    # Build downstream schedules first so IS can reference them.
    build_debt(wb.create_sheet(SH_DEBT))
    build_ppe(wb.create_sheet(SH_PPE))
    build_wc(wb.create_sheet(SH_WC))
    build_tax(wb.create_sheet(SH_TAX))

    # IS — references Tax/Debt/PPE rows now that they exist.
    is_ws = wb.create_sheet(SH_IS)
    build_is_v2(is_ws)
    # Sanity: verify actual IS rows match the pre-registered layout.
    for key, predicted in PRE_IS_LAYOUT.items():
        actual = IS_ROWS.get(key)
        if actual is not None and actual != predicted:
            raise RuntimeError(f"IS layout drift: '{key}' predicted row {predicted}, "
                               f"actual row {actual}. Update PRE_IS_LAYOUT.")

    # Segments — requires IS EBITDA to reconcile (build after IS).
    build_segments(wb.create_sheet(SH_SEG))

    build_equity(wb.create_sheet(SH_EQ))
    build_bs(wb.create_sheet(SH_BS))
    build_cfs(wb.create_sheet(SH_CFS))

    # Patch the BS cash cell that the Debt sheet references.
    # We wrote it as BS!<col>{BS_PLACEHOLDER_CASH}, but at write-time the placeholder
    # was the row offset BEFORE BS was built. Now we need to fix it to BS_ROWS["cash"].
    # We need to rewrite Debt's cash_offset row formulas with the correct cash row.
    debt_ws = wb[SH_DEBT]
    target_cash_row = BS_ROWS["cash"]
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        debt_ws.cell(row=DEBT_ROWS["cash_offset"], column=2 + t,
                     value=f"=-BS!{col}{target_cash_row}").font = formula_font

    # Patch BS cash formula that references CFS!{col}{CFS_PLACEHOLDER_NETCASH}.
    # When BS was built, CFS rows weren't known yet. Now CFS_ROWS["netcash"] is known.
    bs_ws = wb[SH_BS]
    target_netcash_row = CFS_ROWS["netcash"]
    for t in range(1, 6):
        col = get_column_letter(2 + t)
        prev = get_column_letter(2 + t - 1)
        bs_ws.cell(row=BS_ROWS["cash"], column=2 + t,
                   value=f"={prev}{BS_ROWS['cash']}+CFS!{col}{target_netcash_row}").font = formula_font

    build_budget(wb.create_sheet(SH_BUD))
    build_scenarios(wb.create_sheet(SH_SCEN))
    build_sensitivity(wb.create_sheet(SH_SENS))
    build_valuation(wb.create_sheet(SH_VAL))
    build_checks(wb.create_sheet(SH_CHK))

    wb.active = 0
    out = "telecom-model.xlsx"
    wb.save(out)
    print(f"Wrote {out}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
