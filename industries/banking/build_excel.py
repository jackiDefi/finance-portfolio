#!/usr/bin/env python3
"""
Banking — simple 4-sheet operating model (Cover / Inputs / Forecast / Valuation).

A bank P&L (NII + fees -> revenue -> pre-provision profit -> provisions ->
net income), a tangible-equity rollforward, and an equity-based valuation
(justified P/TBV = (ROTE - g)/(CoE - g)) — NOT a DCF, which is the correct
frame for a leveraged spread business.

Convention:  blue = input,  black = formula,  bold = output / subtotal.

Run:  python3 build_excel.py   ->   writes banking-model.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------------------
# Defaults approximate HSBC FY2024
# ----------------------------------------------------------------------------
INPUTS = [
    # (label, value, kind)  kind: "num" | "pct" | "pct2" | "int"
    ("Earning assets (Y0)",        2300.0, "num"),
    ("Asset growth",               0.03,   "pct"),
    ("Net interest margin (NIM)",  0.016,  "pct2"),
    ("Fee & other income (Y0)",    22.0,   "num"),
    ("Cost-to-income ratio",       0.48,   "pct"),
    ("Cost of risk",               0.0020, "pct2"),
    ("Tax rate",                   0.22,   "pct"),
    ("Tangible equity (Y0)",       155.0,  "num"),
    ("Dividend payout",            0.50,   "pct"),
    ("Cost of equity (CoE)",       0.105,  "pct"),
    ("Terminal growth",            0.03,   "pct"),
    ("Shares outstanding (m)",     17800,  "int"),
]
COMPANY = "HSBC Holdings plc"

# ----------------------------------------------------------------------------
NAVY, ACCENT, BLUE, GREY = "1F3A5F", "C8884A", "1155CC", "F2F4F7"
f_title   = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
f_h2      = Font(name="Calibri", size=12, bold=True, color=NAVY)
f_label   = Font(name="Calibri", size=11, color="222222")
f_input   = Font(name="Calibri", size=11, color=BLUE, bold=True)
f_formula = Font(name="Calibri", size=11, color="222222")
f_bold    = Font(name="Calibri", size=11, bold=True, color="111111")
f_unit    = Font(name="Calibri", size=9, italic=True, color="888888")
f_hdr     = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
f_small   = Font(name="Calibri", size=9, color="666666")
fill_title = PatternFill("solid", fgColor=NAVY)
fill_hdr   = PatternFill("solid", fgColor=ACCENT)
fill_grey  = PatternFill("solid", fgColor=GREY)
right  = Alignment(horizontal="right")
center = Alignment(horizontal="center")
left   = Alignment(horizontal="left")

PCT, PCT2, NUM, INT, MULT = "0.0%", "0.00%", "#,##0.0", "#,##0", '0.00"x"'
YEARS = 5
YCOLS = [get_column_letter(3 + t) for t in range(YEARS)]  # C..G


def build(out_path):
    wb = Workbook()

    # ===================== COVER =====================
    ws = wb.active
    ws.title = "Cover"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 92
    ws["B2"] = "Banking — 5-Year Operating Model"
    ws["B2"].font = f_title; ws["B2"].fill = fill_title; ws["B2"].alignment = left
    ws.row_dimensions[2].height = 30
    rows = [
        "", f"Case study calibration: {COMPANY}", "",
        "MODEL MAP",
        "  Cover       — this sheet: map, conventions, disclaimer",
        "  Inputs      — every driver in one place (blue cells are inputs)",
        "  Forecast    — 5-year bank P&L, tangible-equity rollforward, ratios",
        "  Valuation   — justified P/TBV, implied equity value, per-share",
        "",
        "CONVENTIONS",
        "  Blue font  = input you can change",
        "  Black font = formula (do not overwrite)",
        "  Bold       = output / subtotal",
        "",
        "METHOD",
        "  NII = earning assets x NIM.  Revenue = NII + fee income.",
        "  Pre-provision profit = revenue - operating expenses.",
        "  Net income = (pre-provision profit - provisions) x (1 - tax).",
        "  Tangible equity rolls forward by retained earnings.",
        "  ROTE = net income / opening tangible equity.",
        "  Justified P/TBV = (ROTE - g) / (CoE - g)  [equity-based, NOT a DCF].",
        "",
        "DISCLAIMER",
        "  Default values approximate HSBC's most recent reported year.",
        "  Starting positions are illustrative. Replace blue input cells with",
        "  audited figures before any real use. Analytical template only —",
        "  not investment advice.",
    ]
    r = 4
    for line in rows:
        c = ws.cell(row=r, column=2, value=line)
        c.font = f_h2 if (line.isupper() and line) else f_label
        r += 1

    # ===================== INPUTS =====================
    wi = wb.create_sheet("Inputs")
    wi.sheet_view.showGridLines = False
    wi.column_dimensions["A"].width = 32
    wi.column_dimensions["B"].width = 14
    wi.column_dimensions["C"].width = 12
    wi["A1"] = "Inputs — drivers"
    wi["A1"].font = f_title; wi["A1"].fill = fill_title
    wi.merge_cells("A1:C1"); wi.row_dimensions[1].height = 26
    for col, txt in (("A3", "Driver"), ("B3", "Value"), ("C3", "Unit")):
        wi[col] = txt; wi[col].font = f_hdr; wi[col].fill = fill_hdr
    wi["B3"].alignment = center
    row = 4
    for label, value, kind in INPUTS:
        wi.cell(row=row, column=1, value=label).font = f_label
        c = wi.cell(row=row, column=2, value=value); c.font = f_input; c.alignment = right
        c.number_format = {"num": NUM, "pct": PCT, "pct2": PCT2, "int": INT}[kind]
        u = {"num": "USD bn", "pct": "%", "pct2": "%", "int": "millions"}[kind]
        wi.cell(row=row, column=3, value=u).font = f_unit
        row += 1
    IB = {label: f"Inputs!$B${4 + i}" for i, (label, _, _) in enumerate(INPUTS)}
    assets0 = IB["Earning assets (Y0)"]; gr = IB["Asset growth"]
    nim = IB["Net interest margin (NIM)"]; fee0 = IB["Fee & other income (Y0)"]
    cir = IB["Cost-to-income ratio"]; cor = IB["Cost of risk"]
    tax = IB["Tax rate"]; eq0 = IB["Tangible equity (Y0)"]
    payout = IB["Dividend payout"]; coe = IB["Cost of equity (CoE)"]
    tg = IB["Terminal growth"]; sh = IB["Shares outstanding (m)"]

    # ===================== FORECAST =====================
    wf = wb.create_sheet("Forecast")
    wf.sheet_view.showGridLines = False
    wf.column_dimensions["A"].width = 28
    for col in YCOLS:
        wf.column_dimensions[col].width = 12
    wf["A1"] = "Forecast — USD billion"
    wf["A1"].font = f_title; wf["A1"].fill = fill_title
    wf.merge_cells(f"A1:{YCOLS[-1]}1"); wf.row_dimensions[1].height = 26
    wf["A3"] = "USD billion"; wf["A3"].font = f_small
    for t, col in enumerate(YCOLS, start=1):
        c = wf[f"{col}3"]; c.value = f"Y{t}"; c.font = f_hdr; c.fill = fill_hdr; c.alignment = center

    R = {
        "assets": 4, "fee": 5, "nii": 6, "rev": 7, "opex": 8, "ppp": 9,
        "prov": 10, "pbt": 11, "tax": 12, "ni": 13, "eqo": 14, "eqc": 15,
    }
    labels = {
        "assets": "Earning assets (avg)", "fee": "Fee & other income", "nii": "Net interest income",
        "rev": "Total revenue", "opex": "Operating expenses", "ppp": "Pre-provision profit",
        "prov": "Loan-loss provisions", "pbt": "Pre-tax profit", "tax": "Tax", "ni": "Net income",
        "eqo": "Tangible equity (opening)", "eqc": "Tangible equity (closing)",
    }
    bold_rows = {"rev", "ppp", "pbt", "ni"}
    for key, ridx in R.items():
        c = wf.cell(row=ridx, column=1, value=labels[key])
        c.font = f_bold if key in bold_rows else f_label

    for t, col in enumerate(YCOLS, start=1):
        prev = YCOLS[t - 2] if t > 1 else None
        wf[f"{col}{R['assets']}"] = f"={assets0}*(1+{gr})^{t}"
        wf[f"{col}{R['fee']}"]    = f"={fee0}*(1+{gr})^{t}"
        wf[f"{col}{R['nii']}"]    = f"={col}{R['assets']}*{nim}"
        wf[f"{col}{R['rev']}"]    = f"={col}{R['nii']}+{col}{R['fee']}"
        wf[f"{col}{R['opex']}"]   = f"={col}{R['rev']}*{cir}"
        wf[f"{col}{R['ppp']}"]    = f"={col}{R['rev']}-{col}{R['opex']}"
        wf[f"{col}{R['prov']}"]   = f"={col}{R['assets']}*{cor}"
        wf[f"{col}{R['pbt']}"]    = f"={col}{R['ppp']}-{col}{R['prov']}"
        wf[f"{col}{R['tax']}"]    = f"=IF({col}{R['pbt']}>0,{col}{R['pbt']}*{tax},0)"
        wf[f"{col}{R['ni']}"]     = f"={col}{R['pbt']}-{col}{R['tax']}"
        wf[f"{col}{R['eqo']}"]    = (f"={eq0}" if t == 1 else f"={prev}{R['eqc']}")
        wf[f"{col}{R['eqc']}"]    = f"={col}{R['eqo']}+{col}{R['ni']}*(1-{payout})"
        for key, ridx in R.items():
            cell = wf[f"{col}{ridx}"]; cell.number_format = NUM
            cell.font = f_bold if key in bold_rows else f_formula; cell.alignment = right

    rr = 17
    wf.cell(row=rr, column=1, value="Key ratios").font = f_h2
    ratio_defs = [
        ("ROTE", lambda col: f"={col}{R['ni']}/{col}{R['eqo']}", PCT),
        ("Cost-to-income", lambda col: f"={cir}", PCT),
        ("NIM", lambda col: f"={nim}", PCT2),
        ("Cost of risk", lambda col: f"={cor}", PCT2),
    ]
    for i, (lab, fn, fmt) in enumerate(ratio_defs):
        ridx = rr + 1 + i
        wf.cell(row=ridx, column=1, value=lab).font = f_label
        for col in YCOLS:
            c = wf[f"{col}{ridx}"]; c.value = fn(col); c.number_format = fmt
            c.font = f_formula; c.alignment = right

    # ===================== VALUATION =====================
    wv = wb.create_sheet("Valuation")
    wv.sheet_view.showGridLines = False
    wv.column_dimensions["A"].width = 32
    wv.column_dimensions["B"].width = 14
    wv["A1"] = "Valuation — Justified P/TBV"
    wv["A1"].font = f_title; wv["A1"].fill = fill_title
    wv.merge_cells("A1:B1"); wv.row_dimensions[1].height = 26
    wv["A3"] = "Equity-based valuation (not a DCF)"; wv["A3"].font = f_h2

    summary = [
        ("Normalized ROTE (Y1)",      f"=Forecast!C{R['ni']}/Forecast!C{R['eqo']}", PCT,  True),
        ("Cost of equity (CoE)",      f"={coe}",                                     PCT,  False),
        ("Terminal growth (g)",       f"={tg}",                                      PCT,  False),
        ("Justified P/TBV",           "=(B5-B7)/(B6-B7)",                            MULT, True),
        ("Tangible equity Y0 (USD bn)", f"={eq0}",                                   NUM,  False),
        ("Implied equity value (USD bn)", "=B8*B9",                                  NUM,  True),
        ("Shares outstanding (m)",    f"={sh}",                                      INT,  False),
        ("Per-share value (USD)",     "=B10*1000/B11",                               NUM,  True),
    ]
    rstart = 5
    for i, (lab, formula, fmt, is_bold) in enumerate(summary):
        ridx = rstart + i
        a = wv.cell(row=ridx, column=1, value=lab); a.font = f_bold if is_bold else f_label
        b = wv.cell(row=ridx, column=2, value=formula); b.number_format = fmt
        b.font = f_bold if is_bold else f_formula; b.alignment = right
        if is_bold:
            a.fill = fill_grey; b.fill = fill_grey

    wv.cell(row=14, column=1,
            value="Note: a bank has no enterprise value — leverage is the business, so").font = f_small
    wv.cell(row=15, column=1,
            value="EV/EBITDA and operating-cash DCFs do not apply. P/TBV is the frame.").font = f_small

    wb.save(out_path)
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    import os
    here = os.path.dirname(os.path.abspath(__file__))
    build(os.path.join(here, "banking-model.xlsx"))
